#!/usr/bin/python3
# -*- coding: utf-8 -*
import dataclasses
import json
import os.path
import shutil
import sqlite3
import subprocess
import time
import urllib.request
import zipfile
from dataclasses import field, dataclass
from io import BytesIO
from typing import List, Dict, Union

import dbf
import openpyxl
import requests
from lxml import etree
from openpyxl import Workbook

from allplan_manage import AllplanDatas
from hierarchy_qs import *
from main_datas import *
from tools import afficher_message as msg, find_new_title, settings_read, settings_save, read_file_to_list
from tools import convertir_bytes, find_folder_path, validation_fichier_xml, get_value_is_valid
from tools import open_file, find_filename, convertir_nom_fichier_correct, settings_save_value
from tools import application_title
from translation_manage import *
from ui_bcm import Ui_BcmSettings


def a___________________detection______():
    pass


class BddTypeDetection(QObject):

    def __init__(self):
        super().__init__()

        self.error_message = ""
        self.bdd_type = ""
        self.bdd_title = ""
        self.file_path = ""

    def search_bdd_type(self, file_path: str):

        error_txt = self.tr("Erreur")

        bible_txt = self.tr("Bible externe")

        if file_path == "":
            error_message = self.tr("Aucun chemin défini.")

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        if file_path.lower().endswith(".txt"):
            part1 = self.tr("Ce format de fichier n'est pas pris en charge.")
            part2 = self.tr("Si ce fichier est un fichier csv, il est nécessaire de changer l'extention.")

            error_message = f"{part1}\n {part2}"

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        if file_path.startswith("https"):

            self.file_path = file_path

            # ---------------------------------------
            # SYNERMI
            # ---------------------------------------

            if type_synermi.lower() in file_path.lower():
                self.bdd_type = type_synermi
                self.bdd_title = "Synermi"
                return True

            # ---------------------------------------
            # CAPMI
            # ---------------------------------------

            if type_capmi.lower() in file_path.lower():
                self.bdd_type = type_capmi
                self.bdd_title = "Capmi"
                return True

            # ---------------------------------------
            # PROGEMI
            # ---------------------------------------

            if type_progemi.lower() in file_path.lower():
                self.bdd_type = type_progemi
                self.bdd_title = "Progemi"
                return True

            # ---------------------------------------
            # Excel
            # ---------------------------------------

            if ".xlsx" in file_path.lower() and "/" in file_path:
                self.bdd_type = type_excel

                path_split = file_path.split("/")

                for part in path_split:

                    part = part.lower()

                    if ".xlsx" not in part:
                        continue

                    title = part.replace(".xlsx", "").strip()

                    self.bdd_title = title.title()
                    return True

                self.bdd_title = f"{bible_txt} - Excel"
                return True

            self.bdd_type = type_extern
            self.bdd_title = bible_txt

            return True

        title = find_filename(file_path)

        if title == "":
            error_message = self.tr("Une erreur est survenue.")

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        temp = file_path.upper()

        self.file_path = file_path

        error_message = self.tr("Cette base de données n'a pas été reconnue.")

        # ---------------------------------------
        # KUKAT
        # ---------------------------------------

        if temp.endswith(".KAT"):
            self.bdd_type = bdd_type_kukat
            self.bdd_title = f"{bible_txt} - Kukat"
            return True

        # ---------------------------------------
        # EXCEL / CSV
        # ---------------------------------------

        if temp.endswith(".XLSX") or temp.endswith(".CSV"):
            try:
                workbook = openpyxl.load_workbook(self.file_path, data_only=True)
                sheet = workbook.active

                cell_value = sheet['A1'].value

                if cell_value == "Teilleistungsnummer":
                    self.bdd_type = bdd_type_nevaris_xlsx
                    self.bdd_title = title.title()
                    return True

            except Exception:
                pass

            self.bdd_type = type_excel
            self.bdd_title = title.title()
            return True

        # ---------------------------------------
        # MXDB
        # ---------------------------------------

        if temp.endswith(".MXDB"):
            self.bdd_type = type_mxdb
            self.bdd_title = title.title()
            return True

        # ---------------------------------------
        # XML
        # ---------------------------------------

        if temp.endswith(".XML"):

            try:

                tree = etree.parse(file_path)
                root = tree.getroot()

                brand_txt = root.tag

                # ---------------------------------------
                # Nevaris
                # ---------------------------------------

                if brand_txt == '{urn:schemas-microsoft-com:office:spreadsheet}Workbook':

                    try:

                        first_row = root.find(path='.//ss:Row[2]',
                                              namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})

                        if first_row is None:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        style_id = first_row.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}StyleID', '')

                        if not isinstance(style_id, str):
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        if 'NEVARIS_STYLE_H1' != style_id:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        datas = first_row.findall(path=".//ss:Data",
                                                  namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})

                        if len(datas) < 2:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        bdd_type = datas[0].text

                        if bdd_type != "Raumelement":
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        title = datas[1].text

                        if not isinstance(title, str):
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                    except Exception:
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    self.bdd_type = bdd_type_nevaris
                    self.bdd_title = title
                    return True

                # ---------------------------------------
                # SMARTCATALOG
                # ---------------------------------------

                if brand_txt == "AllplanCatalog":

                    check_version = root.find("Node")

                    if check_version is None:
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    self.bdd_type = bdd_type_xml
                    self.bdd_title = title
                    return True

                # ---------------------------------------
                # SMARTCATALOG
                # ---------------------------------------

                check_version = root.find("Folder")

                if check_version is not None:

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt.title()
                    return True

                # ---------------------------------------
                # SMARTCATALOG - Extern
                # ---------------------------------------

                check_version = root.find("Dossier")

                if check_version is not None:

                    # ---------------------------------------
                    # GIMI
                    # --------------------------------------

                    if brand_txt.upper() == "GIMI":

                        self.bdd_type = type_gimi
                        self.bdd_title = "Gimi"
                        return True

                    # ---------------------------------------
                    # Other
                    # --------------------------------------

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt.title()
                    return True

                # ---------------------------------------
                # SMARTCATALOG - Extern - Old
                # ---------------------------------------

                check_version = root.find("Dossier")

                if check_version is not None:

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt
                    return True

                self.error_message = f"{error_txt} -- {error_message}"

                return False

            except Exception:

                self.error_message = f"{error_txt} -- {error_message}"

            return False

        # ---------------------------------------
        # ARA
        # ---------------------------------------

        if temp.endswith(".ARA"):

            file_path = unpack_ara_file(file_path=file_path)

            if file_path == "":
                return False

            self.file_path = file_path

            temp = file_path.upper()

        # ---------------------------------------
        # DBF
        # ---------------------------------------
        if temp.endswith(".DBF"):

            try:

                with dbf.Table(filename=file_path) as table:

                    if len(table) == 0:
                        self.error_message = f"{error_txt} -- {error_message}"

                        return False

                    liste_colonnes = table.field_names

                    if "VWKTX" not in liste_colonnes:
                        self.error_message = f"{error_txt} -- {error_message}"

                        return False

                    record = table.first_record
                    item_type: str = convertir_bytes(record.VWTPU)
                    code_item_type: str = convertir_bytes(record.VWTYP)
                    title = convertir_bytes(record.VWKTX)

                    title = convertir_nom_fichier_correct(title)

                    if code_item_type != "X":
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    # ---------------------------------------
                    # ALLMETRE - EURICIEL
                    # ---------------------------------------

                    if item_type == "Projekt":
                        self.bdd_type = type_allmetre_e
                        self.bdd_title = title.title()
                        return True

                    # ---------------------------------------
                    # ALLMETRE - AJSOFT
                    # ---------------------------------------

                    if item_type == "Pos.":
                        self.bdd_type = type_allmetre_a
                        self.bdd_title = title.title()
                        return True

                    for record in table:

                        if dbf.is_deleted(record):
                            continue

                        code_item_type: str = convertir_bytes(record.VWTYP)

                        # ---------------------------------------
                        # BCM - MATERIAL
                        # ---------------------------------------

                        if code_item_type == "E":
                            self.bdd_type = bdd_type_bcm
                            self.bdd_title = title.title()
                            return True

                        # ---------------------------------------
                        # BCM - COMPONENT
                        # ---------------------------------------

                        if code_item_type == "L":
                            self.bdd_type = type_bcm_c
                            self.bdd_title = title.title()
                            return True

            except dbf.DbfError:
                self.error_message = f"{error_txt} -- {error_message}"

            return False

        if temp.endswith(".FIC") or temp.endswith(".NDX") or temp.endswith(".MMO"):

            if title.upper() in ["CM", "ST", "ARTICLES", "PARA"]:

                path_gimi = os.path.dirname(file_path)

                path_gimi = path_gimi.replace("/", "\\")

                if not path_gimi.endswith("\\"):
                    path_gimi += "\\"

                settings_save_value(library_setting_file, "path_gimi", path_gimi)

                # ---------------------------------------
                # GIMI
                # ---------------------------------------

                self.bdd_type = type_gimi
                self.bdd_title = f"{bible_txt} - {type_gimi}"
                return True

            return False

        dict_favoris_allplan = get_favorites_allplan_dict()

        for extension, nom_favoris in dict_favoris_allplan.items():

            if file_path.endswith(extension):
                self.bdd_type = bdd_type_fav
                self.bdd_title = title.title()
                return True

        self.error_message = f"{error_txt} -- {error_message}"
        return False


def unpack_ara_file(file_path: str) -> str:
    # -------------------------
    # Verification file exists
    # -------------------------

    if not isinstance(file_path, str):
        print(f"conver_manage -- unpack_ara_file -- not isinstance(file_path, str)")
        return ""

    if not os.path.exists(file_path):
        print(f"conver_manage -- unpack_ara_file -- not os.path.exists(file_path)")
        return ""

    # -------------------------
    # search folder path
    # -------------------------

    folder_path = find_folder_path(file_path)

    if folder_path == "":
        print(f"conver_manage -- unpack_ara_file -- folder_path == empty")
        return ""

    # -------------------------
    # search filename
    # -------------------------

    file_name = find_filename(file_path=file_path)

    file_name = file_name.strip()

    if file_name == "":
        print(f"conver_manage -- unpack_ara_file -- file_name == empty")
        return ""

    # -------------------------
    # define others paths
    # -------------------------

    export_path = f"{asc_export_path}{file_name}\\"

    json_file_path = f"{export_path}{file_name}.ini"

    export_path_bak = f"{asc_export_path}{file_name}_bak"

    dbf_path = f"{export_path}VW1.DBF"

    # -------------------------
    # search size ARA file
    # -------------------------

    try:
        file_size = os.path.getsize(file_path)

    except OSError as error:
        print(f"conver_manage -- unpack_ara_file -- error delete : {error}")
        return ""

    # -------------------------
    # search old size
    # -------------------------

    file_size_old = 0

    if os.path.exists(json_file_path):

        try:

            with open(json_file_path, 'r', encoding="Utf-8") as file:

                file_size_old = json.load(file)

        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error json read : {error}")
            file_size_old = 0
            pass

    # -------------------------
    # current ARA file and old ARA file are same
    # -------------------------

    if file_size == file_size_old and os.path.exists(dbf_path):
        return dbf_path

    # -------------------------
    # Delete backup folder
    # -------------------------

    if os.path.exists(export_path_bak):

        try:
            shutil.rmtree(export_path_bak)
        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error delete : {error}")
            return ""

    # -------------------------
    # Rename current folder to backup folder
    # -------------------------

    if os.path.exists(export_path):

        try:
            os.rename(export_path, export_path_bak)
        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error rename : {error}")
            return ""

    # -------------------------
    # Create new folder
    # -------------------------

    try:
        os.makedirs(export_path)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error make dir : {error}")
        return ""

    # -------------------------
    # Unzip
    # -------------------------

    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(export_path)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error zip : {error}")
        return ""

    # -------------------------
    # search dbf file
    # -------------------------

    if not os.path.exists(dbf_path):
        print(f"conver_manage -- unpack_ara_file -- not os.path.exists(dbf_path)")
        return ""

    # -------------------------
    # Write json file
    # -------------------------

    try:
        with open(json_file_path, 'w', encoding="Utf-8") as file:

            json.dump(file_size, file, ensure_ascii=False, indent=2)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error json write : {error}")
        return "False"

    # -------------------------
    # Unpack is ok
    # -------------------------

    return dbf_path


def a___________________bcm_component______():
    pass


class ConvertBcmComposants(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        # Projet = 		X
        # Cout = 		K
        # Din276 = 		D
        # Remarque = 	B
        # Répertoire = 	L
        # Titre = 		T
        # Composant =	P
        # Total =		S

        self.allplan = allplan

        self.creation = self.allplan.creation
        self.chemin_fichier = file_path

        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.prices_dict = dict()

    def run(self):

        col_type = 0
        col_desc = 1
        col_code = 2
        col_unite = 3
        col_txt_long = 4
        col_formule_metre = 5
        col_materiaux_dyn = 6
        col_quantite = 7
        col_objet = 8
        col_price = 9

        tps = time.perf_counter()

        self.get_all_prices()

        try:
            with dbf.Table(filename=self.chemin_fichier) as table:

                liste_bdd = list()
                datas_dict = dict()

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    type_ele: str = convertir_bytes(record.VWTYP)

                    if type_ele == "X" or type_ele == "B" or type_ele == "K" or type_ele == "D":
                        continue

                    id_ele_int: str = convertir_bytes(record.VWSRT)

                    if id_ele_int is None:
                        continue

                    id_ele = f"{id_ele_int:010d}"
                    description: str = convertir_bytes(record.VWKTX)

                    if type_ele == "S":
                        liste_bdd.append(id_ele)
                        datas_dict[id_ele] = [type_ele, description]
                        continue

                    code: str = convertir_bytes(record.VWCTX)

                    if code == "":
                        code: str = convertir_bytes(record.VWPNR)

                    if type_ele == "L" or type_ele == "T":
                        liste_bdd.append(id_ele)
                        datas_dict[id_ele] = [type_ele, description, code]
                        continue

                    if type_ele == "P":
                        unite: str = convertir_bytes(record.VWDIM)
                        txt_long: str = convertir_bytes(record.VWLTX)
                        formule_b: str = convertir_bytes(record.VWATR)
                        formule_metre, materiaux_dyn, objet, quantite = self.allplan.convertir_formule_bdd(formule_b)

                        cid: str = convertir_bytes(record.VWCID)

                        price = self.prices_dict.get(cid, "")

                        liste_bdd.append(id_ele)

                        datas_dict[id_ele] = [type_ele, description, code, unite, txt_long,
                                              formule_metre, materiaux_dyn, objet, quantite, price]

        except Exception as erreur:

            print(f"convert_manage -- ConvertBcmComposants -- run -- Erreur : {erreur}")

            self.loading_completed.emit(self.model, list(), list())
            return

        liste_bdd.sort()
        parent = self.model.invisibleRootItem()

        for id_art, id_ele in enumerate(liste_bdd):

            type_ele = datas_dict[id_ele][col_type]

            if type_ele == "L":
                parent = self.model.invisibleRootItem()

            elif type_ele == "S":

                ancien_parent = parent.parent()

                if ancien_parent is None:

                    parent = self.model.invisibleRootItem()

                else:
                    parent = ancien_parent

                continue

            code = datas_dict[id_ele][col_code]
            description = datas_dict[id_ele][col_desc]

            if type_ele == "L" or type_ele == "T":
                liste_ajouter = self.creation.folder_line(value=code,
                                                          description=description,
                                                          icon_path="",
                                                          tooltips=False)

                parent.appendRow(liste_ajouter)

                parent: MyQstandardItem = liste_ajouter[col_cat_value]

                continue

            if type_ele == "P":

                unite: str = datas_dict[id_ele][col_unite]
                texte_long: str = datas_dict[id_ele][col_txt_long]
                formule_metre: str = datas_dict[id_ele][col_formule_metre]
                materiaux_dyn: str = datas_dict[id_ele][col_materiaux_dyn]
                quantite: str = datas_dict[id_ele][col_quantite]
                objet: str = datas_dict[id_ele][col_objet]
                price: str = datas_dict[id_ele][col_price]

                liste_ajouter = self.creation.component_line(value=code, description=description)

                parent.appendRow(liste_ajouter)

                composant: MyQstandardItem = liste_ajouter[col_cat_value]

                if unite != "":
                    liste_ajouter = self.creation.attribute_line(value=unite, number="202")
                    composant.appendRow(liste_ajouter)

                if texte_long != "":
                    liste_ajouter = self.creation.attribute_line(value=texte_long, number="208")
                    composant.appendRow(liste_ajouter)

                if formule_metre != "":
                    liste_ajouter = self.creation.attribute_line(value=formule_metre, number="267")
                    composant.appendRow(liste_ajouter)

                if materiaux_dyn != "":
                    liste_ajouter = self.creation.attribute_line(value=materiaux_dyn, number="96")
                    composant.appendRow(liste_ajouter)

                if objet != "":
                    liste_ajouter = self.creation.attribute_line(value=objet, number="76")
                    composant.appendRow(liste_ajouter)

                if quantite != "":
                    liste_ajouter = self.creation.attribute_line(value=quantite, number="215")
                    composant.appendRow(liste_ajouter)

                if price != "":
                    liste_ajouter = self.creation.attribute_line(value=price, number="203")
                    composant.appendRow(liste_ajouter)

        print(f"ConvertBcmComposants : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())

    def get_all_prices(self):

        file_name = find_filename(file_path=self.chemin_fichier)

        file_name_new = file_name.lower().replace("vw", "pr")

        folder_path = find_folder_path(file_path=self.chemin_fichier)

        file_path = f"{folder_path}{file_name_new}.dbf"

        if not os.path.exists(file_path):
            return

        tps = time.perf_counter()

        try:
            with dbf.Table(filename=file_path) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    price_float = convertir_bytes(record.PREPB)

                    if not isinstance(price_float, float):
                        continue

                    if price_float == 0:
                        continue

                    price_str = f"{price_float:.2f}"

                    current_code = convertir_bytes(record.PRIDPOS)

                    if not isinstance(current_code, str):
                        continue

                    if current_code in self.prices_dict:
                        continue

                    self.prices_dict[current_code] = price_str

        except Exception as error:
            print(f"convert_manage -- ConvertBcmComposants -- get_all_prices -- Erreur : {error}")
            return

        print(f"ConvertBcmComposants -- Price : {time.perf_counter() - tps}s")


class BcmArticle:
    id_ele = ""
    index_liste = 0
    type_ele = ""
    code = ""
    description = ""
    niveau = 0
    txt_long = ""
    unite = ""
    formule = ""
    materiaux_dyn = ""
    objet = ""
    quantite = ""
    price = ""

    son_parent_article = None
    son_parent_article_type_ele = None

    liste_enfants = list()
    liste_enfants_type_ele = list()

    dossier_a_creer = False
    dossier = False


def a___________________bcm_material______():
    pass


class ConvertBcmOuvrages(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)
    errors_signal = pyqtSignal(list)

    def __init__(self, allplan, file_path: str, bdd_title: str, conversion=False):
        super().__init__()

        # E = Ouvrage
        # P = Composant
        # V = Lien
        # S = Total

        self.allplan = allplan

        if conversion:
            self.allplan.creation.attributes_datas.clear()

        self.creation = self.allplan.creation

        self.chemin_fichier = file_path

        self.conversion = conversion

        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.liste_table = list()
        self.liste_niveau_zero = list()
        self.dict_ouvrages = dict()
        self.liste_description_ouvrage = list()
        self.liste_code_ouvrage = list()
        self.liste_erreurs = list()

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

        self.liens_dict = dict()
        self.prices_dict = dict()

    def run(self):

        self.get_all_prices()

        if not self.bcm_chargement_bdd():
            self.loading_completed.emit(self.model, list(), list())
            return

        tps = time.perf_counter()

        self.bcm_treeview_creation()

        print(f"ConvertBcmOuvarges : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())
        return

    @staticmethod
    def a___________________bdd_bcm_analyse__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def bcm_chargement_bdd(self):

        try:
            with dbf.Table(filename=self.chemin_fichier) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    code_type_ele: str = convertir_bytes(record.VWTYP)

                    if code_type_ele != "E" and code_type_ele != "P" and code_type_ele != "S" and code_type_ele != "V":
                        continue

                    id_ele = convertir_bytes(record.VWSRT)

                    nouvel_article = BcmArticle()
                    nouvel_article.id_ele = id_ele
                    nouvel_article.type_ele = code_type_ele

                    # Gestion code
                    code = convertir_bytes(record.VWCTX)

                    if code == "":
                        code = convertir_bytes(record.VWPNR)

                    if code == "":
                        code = "sans_code"

                    cid: str = convertir_bytes(record.VWCID)

                    price = self.prices_dict.get(cid, "")

                    nouvel_article.code = code

                    nouvel_article.description = convertir_bytes(record.VWKTX)
                    nouvel_article.niveau = convertir_bytes(record.VWHIR)
                    nouvel_article.txt_long = convertir_bytes(record.VWLTX)

                    # Gestion unités
                    unites_tps = convertir_bytes(record.VWDIM)
                    nouvel_article.unite = self.allplan.convert_unit(unites_tps)

                    nouvel_article.price = price

                    # Gestion formules
                    formule_tps = convertir_bytes(record.VWATR)

                    formule_metre, materiaux_dyn, objet, quantite = self.allplan.convertir_formule_bdd(formule_tps)

                    nouvel_article.formule = formule_metre
                    nouvel_article.materiaux_dyn = materiaux_dyn
                    nouvel_article.objet = objet
                    nouvel_article.quantite = quantite

                    nouvel_article.son_parent_article = None
                    nouvel_article.son_parent_article_type_ele = None

                    nouvel_article.liste_enfants = list()
                    nouvel_article.liste_enfants_type_ele = list()

                    self.liste_table.append([id_ele, nouvel_article])

        except Exception as erreur:

            print(f"widget_onglet -- QThreadBddBcmOuvrages -- Erreur : {erreur}")

            self.loading_completed.emit(self.model, list(), list())
            return False

        if len(self.liste_table) == 0:
            print(f"widget_onglet -- QThreadBddBcmOuvrages -- Erreur : liste_table == 0")

            self.loading_completed.emit(self.model, list(), list())
            return False

        self.liste_table.sort(key=lambda x: int(x[0]))

        nb_elements_liste_table = len(self.liste_table)
        nb_index_liste_table = nb_elements_liste_table - 1

        for index_ele_actuel, liste_donnees_actuel in enumerate(self.liste_table):

            # Mise ene mémoire de l'index suivant
            index_ele_suivant = index_ele_actuel + 1

            # Définition de l'article analysé
            liste_donnees_actuel: list

            article_actuel: BcmArticle = liste_donnees_actuel[1]

            # Ajout de l'index de la liste_table dans l'article analysé
            article_actuel.index_liste = index_ele_actuel

            # Mise en mémoire du niveau de l'article analysé
            niveau_actuel = article_actuel.niveau
            type_ele_actuel = article_actuel.type_ele

            # Mise en mémoire du niveau des enfants de cet article analysé
            niveau_enfants = niveau_actuel + 1

            # -----------------------------------------
            # Ajout des enfants dans l'article analysé
            # -----------------------------------------

            if type_ele_actuel == "E":

                if article_actuel.code not in self.dict_ouvrages:
                    self.dict_ouvrages[article_actuel.code] = article_actuel
                # else:
                #     print(f"bdd_chargement_bdd -- {articles_actuel.code} déjà dans self.dict_ouvrages")

            if index_ele_suivant <= nb_index_liste_table:

                if type_ele_actuel == "E":
                    self.bcm_recherche_enfants(index_ele_suivant=index_ele_suivant,
                                               nb_elements_liste_table=nb_elements_liste_table,
                                               niveau_enfants=niveau_enfants,
                                               article_actuel=article_actuel)
            if niveau_actuel == 1:
                self.liste_niveau_zero.append(article_actuel)

        for index_ele_actuel, liste_donnees_actuel in enumerate(self.liste_table):

            article_actuel: BcmArticle = liste_donnees_actuel[1]
            type_ele_actuel = article_actuel.type_ele

            if type_ele_actuel != "V":
                continue

            self.bcm_chargement_bdd_creation_lien(article_actuel.code)

        return True

    def bcm_chargement_bdd_creation_lien(self, nom_lien: str):

        if nom_lien not in self.dict_ouvrages:
            print(f"widget_onglet -- QThreadBddBcmOuvrages -- bcm_chargement_bdd_creation_lien --> "
                  f" {nom_lien} not in self.dict_ouvrages")
            return

        if nom_lien in self.liens_dict:
            return

        article_lien: BcmArticle = self.dict_ouvrages[nom_lien]

        liste_enfants_lien: list = article_lien.liste_enfants

        for enfant_lien in liste_enfants_lien:

            enfant_lien: BcmArticle

            enfant_lien_type_ele = enfant_lien.type_ele
            enfant_lien_code = enfant_lien.code

            if enfant_lien_type_ele == "P":

                self.bcm_chargement_bdd_ajouter_composant_dans_lien(nom_lien, enfant_lien)

            elif enfant_lien_type_ele == "E":

                if nom_lien == enfant_lien_code:
                    return

                self.bcm_chargement_bdd_creation_lien(enfant_lien_code)

                liste_composants = self.liens_dict[enfant_lien_code]

                for composant in liste_composants:
                    composant: BcmArticle

                    self.bcm_chargement_bdd_ajouter_composant_dans_lien(nom_lien, composant)

            elif enfant_lien_type_ele == "V":

                if enfant_lien_code not in self.dict_ouvrages:
                    print(f"widget_onglet -- QThreadBddBcmOuvrages -- bcm_chargement_bdd_creation_lien --> "
                          f"{enfant_lien_code} not in self.dict_ouvrages")
                    continue

                if enfant_lien_code not in self.liens_dict:
                    self.bcm_chargement_bdd_creation_lien(enfant_lien_code)

                liste_composants = self.liens_dict[enfant_lien_code]

                for composant in liste_composants:
                    composant: BcmArticle

                    self.bcm_chargement_bdd_ajouter_composant_dans_lien(nom_lien, composant)

    def bcm_chargement_bdd_ajouter_composant_dans_lien(self, nom_lien: str, composant: BcmArticle):
        """

        :param nom_lien:
        :param composant:
        :return:
        """

        if nom_lien not in self.liens_dict:
            self.liens_dict[nom_lien] = list()

        liste_actuel: list = self.liens_dict[nom_lien]

        if composant in liste_actuel:
            return

        liste_actuel.append(composant)

        # print(f"Ouvrage : {nom_lien} -- Ajout Composant : {composant.code}")

    def bcm_recherche_enfants(self, index_ele_suivant: int, nb_elements_liste_table: int, niveau_enfants: int,
                              article_actuel: BcmArticle):

        if article_actuel.type_ele == "V":

            if article_actuel.son_parent_article is None:
                print("widget_onglet -- QThreadBddBcmOuvrages -- bcm_recherche_enfants --> "
                      f"{article_actuel.description} n'a pas de parent")
            return

        # Parcourir les enfants de l'article analysé jusqu'à son total
        for index_enfant in range(index_ele_suivant, nb_elements_liste_table):

            # Mise en mémoire de l'article enfant
            liste_donnees_enfant_analyser: list = self.liste_table[index_enfant]

            articles_enfant_analyser: BcmArticle = liste_donnees_enfant_analyser[1]

            type_ele_enfant_analyser = articles_enfant_analyser.type_ele
            niveau_enfant_analyser = articles_enfant_analyser.niveau

            # Si le niveau de l'enfant analysé = niveau recherché et que le type d'élément == "Total" ==> retourne list
            if type_ele_enfant_analyser == "S" and niveau_enfant_analyser == niveau_enfants:
                return

                # Si le niveau de l'enfant analysé = niveau recherché == ajout de l'enfant dans la liste
            elif niveau_enfant_analyser == niveau_enfants:
                article_actuel.liste_enfants.append(articles_enfant_analyser)
                article_actuel.liste_enfants_type_ele.append(articles_enfant_analyser.type_ele)

                if articles_enfant_analyser.son_parent_article is None:
                    articles_enfant_analyser.son_parent_article = article_actuel
                    articles_enfant_analyser.son_parent_article_type_ele = article_actuel.type_ele
        return

    def bcm_analyser_avant_ajout_ouvrages(self, article_actuel: BcmArticle):

        liste_enfants_actuel = article_actuel.liste_enfants
        liste_enfants_actuel_type_ele = article_actuel.liste_enfants_type_ele

        if "E" not in liste_enfants_actuel_type_ele:
            return

        for article_enfant in liste_enfants_actuel:

            article_enfant: BcmArticle

            liste_enfants_analyser_type_ele = article_enfant.liste_enfants_type_ele

            if "E" not in liste_enfants_analyser_type_ele:
                continue

            article_actuel.dossier_a_creer = True
            article_enfant.dossier = True

            self.bcm_analyser_avant_ajout_ouvrages(article_enfant)

        return

    def bcm_renommer_ouvrage(self, article_actuel: BcmArticle):
        """
        Permet de renommer un code déjà existant
        :param article_actuel: code actuel
        :return: code renommer
        """

        code_actuel = article_actuel.code

        index_code = 1

        while True:

            new_code_actuel = f"{code_actuel.upper()} - {index_code}"

            if new_code_actuel not in self.liste_code_ouvrage:
                self.liste_code_ouvrage.append(new_code_actuel)

                parent_article = article_actuel.son_parent_article

                if parent_article is None:
                    parent_nom = self.tr("Inconnu")
                else:
                    parent_article: BcmArticle
                    parent_nom = parent_article.description

                a = self.tr("L'ouvrage")
                b = self.tr("existe déjà")
                c = self.tr("et a été renommé en")
                d = self.tr("Parent")

                self.liste_erreurs.append(f"{a} : {code_actuel} {b} "
                                          f"{c} {new_code_actuel}"
                                          f"({d} == {parent_nom})")
                return new_code_actuel

            index_code += 1

            if index_code >= 200:
                e = self.tr("Le nombre maxi de recherche de nouveau nom a été dépassé pour l'ouvrage")

                msg(titre=application_title,
                    message=f"{e} {code_actuel}")
                return ""

    @staticmethod
    def a___________________bdd_bcm_creation__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def bcm_treeview_creation(self):

        hierarchie_parent = self.model.invisibleRootItem()

        for index_article, article_actuel in enumerate(self.liste_niveau_zero):

            article_actuel: BcmArticle

            # code_actuel: str = article_actuel.code
            description_actuel: str = article_actuel.description
            code_actuel: str = article_actuel.code
            liste_enfants: list = article_actuel.liste_enfants
            liste_enfants_type_ele: list = article_actuel.liste_enfants_type_ele

            self.bcm_analyser_avant_ajout_ouvrages(article_actuel)

            dossier_a_creer = article_actuel.dossier_a_creer

            if ("P" in liste_enfants_type_ele or "V" in liste_enfants_type_ele) and \
                    "E" not in liste_enfants_type_ele:

                # 001 ajouter dossier (mystandarditem_dossier) dans hierarchie_parent

                mystandarditem_dossier: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                          code_actuel,
                                                                          description_actuel)

                if description_actuel not in self.liste_description_ouvrage:
                    self.liste_description_ouvrage.append(description_actuel)

                if code_actuel.upper() not in self.liste_code_ouvrage:
                    self.liste_code_ouvrage.append(code_actuel.upper())

                else:

                    code_actuel_tmp = self.bcm_renommer_ouvrage(article_actuel)

                    if code_actuel_tmp == "":
                        print("BcmConvertirBdd -- bcm_treeview_creation -- "
                              "Erreur lors de la recherche d'un nouveau code pour un ouvrage déjà existant : "
                              f"{code_actuel}")

                        continue

                    article_actuel.code = code_actuel_tmp

                # 002 ajouter ouvrage (mystandarditem_actuel) dans mystandarditem_dossier
                mystandarditem_actuel: MyQstandardItem = self.material_add(mystandarditem_dossier, article_actuel)

            else:

                if description_actuel not in self.liste_description_ouvrage:
                    self.liste_description_ouvrage.append(description_actuel)

                if code_actuel.upper() not in self.liste_code_ouvrage:
                    self.liste_code_ouvrage.append(code_actuel.upper())

                else:

                    code_actuel_tmp = self.bcm_renommer_ouvrage(article_actuel)

                    if code_actuel_tmp == "":
                        print("BcmConvertirBdd -- bcm_treeview_creation -- "
                              "Erreur lors de la recherche d'un nouveau code pour un ouvrage déjà existant : "
                              f"{code_actuel}")

                        continue

                    article_actuel.code = code_actuel_tmp

                if dossier_a_creer:

                    # 008 ajouter dossier (mystandarditem_actuel) dans hierarchie_parent
                    mystandarditem_actuel: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                             code_actuel,
                                                                             description_actuel)

                    # 006 ajouter dossier (mystandarditem_dossier) dans mystandarditem_actuel
                    mystandarditem_dossier: MyQstandardItem = self.folder_add(mystandarditem_actuel,
                                                                              code_actuel,
                                                                              description_actuel)

                    # 005 ajouter ouvrage (mystandarditem_ouvrage) dans mystandarditem_dossier
                    mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandarditem_dossier,
                                                                                article_actuel)

                else:

                    # 008 ajouter dossier (mystandarditem_actuel) dans hierarchie_parent
                    mystandarditem_actuel: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                             code_actuel,
                                                                             description_actuel)

                    # 007 ajouter ouvrage (mystandarditem_dossier) dans mystandarditem_actuel
                    mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandarditem_actuel,
                                                                                article_actuel)

                # Ajout de l'ensemble des composants
                self.bcm_treeview_charger_tous_enfants(article_actuel, mystandarditem_ouvrage)

            for enfant in liste_enfants:
                self.bcm_treeview_charger_enfants(enfant, mystandarditem_actuel, dossier_a_creer)

    def bcm_treeview_charger_enfants(self, article_actuel: BcmArticle, hierarchie_parent: MyQstandardItem,
                                     dossier_a_creer):

        type_ele_actuel: str = article_actuel.type_ele
        code_actuel: str = article_actuel.code
        description_actuel: str = article_actuel.description
        liste_enfants: list = article_actuel.liste_enfants
        dossier_a_creer_actuel = article_actuel.dossier_a_creer
        dossier_actuel = article_actuel.dossier

        a = self.tr("Le Composant")
        b = self.tr("a été ignoré")
        c = self.tr("Son parent")
        d = self.tr("n'est pas un ouvrage")
        e = self.tr("Le lien")
        f = self.tr("L'ouvrage")
        g = self.tr("n'existe pas")

        if type_ele_actuel == "E":

            if description_actuel not in self.liste_description_ouvrage:
                self.liste_description_ouvrage.append(description_actuel)

            if code_actuel.upper() not in self.liste_code_ouvrage:
                self.liste_code_ouvrage.append(code_actuel.upper())

            else:

                code_actuel_tmp = self.bcm_renommer_ouvrage(article_actuel)

                if code_actuel_tmp == "":
                    print("BcmConvertirBdd -- bcm_treeview_charger_enfants -- "
                          "Erreur lors de la recherche d'un nouveau code pour un ouvrage déjà existant : "
                          f"{code_actuel}")

                    return

                article_actuel.code = code_actuel_tmp

            if dossier_a_creer and not dossier_actuel:
                # 001 ajouter dossier (mystandard_dossier) dans hierarchie_parent
                mystandard_dossier: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                      code_actuel,
                                                                      description_actuel)

                hierarchie_parent = mystandard_dossier

            if dossier_actuel:

                # 002 ajouter dossier (mystandardwidget_actuel) dans hierarchie_parent
                mystandardwidget_actuel: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                           code_actuel,
                                                                           description_actuel)

                if len(liste_enfants) == 0:

                    # 003 ajouter dossier (mystandardwidget_ouvrage) dans mystandardwidget_actuel
                    mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandardwidget_actuel,
                                                                                article_actuel)

                    # Ajout de l'ensemble des composants si dossier actuel = True
                    self.bcm_treeview_charger_tous_enfants(article_actuel, mystandarditem_ouvrage)

                else:

                    creation_dossier = False

                    for enfant_analyser in liste_enfants:

                        if enfant_analyser.dossier:
                            creation_dossier = True
                            break

                    if not creation_dossier:

                        # 004 ajouter ouvrage (mystandardwidget_ouvrage) dans mystandardwidget_actuel
                        mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandardwidget_actuel,
                                                                                    article_actuel)

                        # Ajout de l'ensemble des composants si dossier actuel = True
                        self.bcm_treeview_charger_tous_enfants(article_actuel, mystandarditem_ouvrage)

                    else:

                        # 005 ajouter dossier (mystandard_dossier) dans mystandardwidget_actuel
                        mystandard_dossier: MyQstandardItem = self.folder_add(mystandardwidget_actuel,
                                                                              code_actuel,
                                                                              description_actuel)

                        # 004 ajouter ouvrage (mystandardwidget_ouvrage) dans mystandardwidget_actuel
                        mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandard_dossier,
                                                                                    article_actuel)

                        # Ajout de l'ensemble des composants si dossier actuel = True
                        self.bcm_treeview_charger_tous_enfants(article_actuel, mystandarditem_ouvrage)

            else:

                # 007 ajouter ouvrage (mystandardwidget_actuel) dans hierarchie_parent
                mystandardwidget_actuel: MyQstandardItem = self.material_add(hierarchie_parent,
                                                                             article_actuel)

            for enfant in liste_enfants:
                enfant: BcmArticle
                self.bcm_treeview_charger_enfants(enfant, mystandardwidget_actuel, dossier_a_creer_actuel)

        elif type_ele_actuel == "P":

            article_parent: BcmArticle = article_actuel.son_parent_article
            liste_type_enfants_du_parent = article_parent.liste_enfants_type_ele

            if "E" in liste_type_enfants_du_parent:
                texte = f"{a} : {description_actuel} {b} " \
                        f"({c} : {article_parent.code} {d})"

                self.liste_erreurs.append(texte)
                return

            # 100 ajouter composant dans hierarchie_parent
            self.component_add(hierarchie_parent, article_actuel)

        elif type_ele_actuel == "V":

            article_parent: BcmArticle = article_actuel.son_parent_article
            liste_type_enfants_du_parent = article_parent.liste_enfants_type_ele

            if "E" in liste_type_enfants_du_parent:
                texte = f"{e} : {description_actuel} {b}" \
                        f"({c} : {article_parent.code} {d})"

                self.liste_erreurs.append(texte)

                return

            if code_actuel not in self.dict_ouvrages:
                texte = f"{e} : {description_actuel} {b} ({f} {article_parent.code} {g})"

                self.liste_erreurs.append(texte)

                return

            if code_actuel not in self.liens_dict:
                texte = f"{e} : {description_actuel} {b}  ({f} {article_parent.code} {g})"

                self.liste_erreurs.append(texte)

                return

            liste_composants = self.liens_dict[code_actuel]

            for article_composant in liste_composants:
                # 101 ajouter composant dans hierarchie_parent
                self.component_add(his_parent=hierarchie_parent, article=article_composant)

    def bcm_treeview_charger_tous_enfants(self, article_actuel: BcmArticle, hierarchie_parent: MyQstandardItem):

        liste_enfants: list = article_actuel.liste_enfants

        for enfant in liste_enfants:

            enfant: BcmArticle
            enfant_type_ele_actuel: str = enfant.type_ele

            if enfant_type_ele_actuel == "P":

                self.component_add(his_parent=hierarchie_parent, article=enfant)

            elif enfant_type_ele_actuel == "E":

                self.bcm_treeview_charger_tous_enfants(enfant, hierarchie_parent)

            elif enfant_type_ele_actuel == "V":

                if enfant.code not in self.dict_ouvrages:
                    print(f"BcmConvertirBdd -- bcm_treeview_charger_tous_enfants -- "
                          f"{enfant.code} not in self.dict_ouvrages")
                    continue

                if enfant.code not in self.liens_dict:
                    print(f"BcmConvertirBdd -- bcm_hierarchie_charger_tous_enfants -- "f""
                          f"{enfant.code} not in liens_dict.dict_ouvrages")

                    continue

                liste_composants = self.liens_dict[enfant.code]

                for article_composant in liste_composants:
                    article_composant: BcmArticle

                    # 102 ajouter composant dans hierarchie_parent
                    self.component_add(his_parent=hierarchie_parent, article=article_composant)

        # print("bcm_treeview_charger_tous_enfants -- fin")

    @staticmethod
    def a___________________bdd_bcm_objets__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def folder_add(self, his_parent: MyQstandardItem, code: str, description: str) -> MyQstandardItem:

        if code.startswith(self.tr("code vide")):
            qs_list = self.creation.folder_line(value=description,
                                                tooltips=False)
        else:
            qs_list = self.creation.folder_line(value=code,
                                                description=description,
                                                tooltips=False)

        his_parent.appendRow(qs_list)
        return qs_list[col_cat_value]

    def material_add(self, his_parent: MyQstandardItem, article: BcmArticle) -> MyQstandardItem:

        qs_list = self.creation.material_line(value=f"{article.code}",
                                              description=article.description,
                                              tooltips=False)

        self.material_list.append(article.code)
        self.material_upper_list.append(article.code.upper())

        qs: MyQstandardItem = qs_list[col_cat_value]

        self.attributes_add(qs=qs, article=article)

        his_parent.appendRow(qs_list)
        return qs

    def component_add(self, his_parent: MyQstandardItem, article: BcmArticle) -> MyQstandardItem:

        qs_list = self.creation.component_line(value=f"{article.code}",
                                               description=article.description,
                                               tooltips=False)

        qs: MyQstandardItem = qs_list[col_cat_value]

        self.attributes_add(qs=qs, article=article)

        his_parent.appendRow(qs_list)
        return qs

    def attributes_add(self, qs: MyQstandardItem, article: BcmArticle):

        if article.txt_long != "":
            attributes_list = self.creation.attribute_line(value=article.txt_long, number="208")
            qs.appendRow(attributes_list)

        if article.unite != "":
            attributes_list = self.creation.attribute_line(value=article.unite, number="202")
            qs.appendRow(attributes_list)

        if article.formule != "":
            attributes_list = self.creation.attribute_line(value=article.formule, number="267")
            qs.appendRow(attributes_list)

        if article.materiaux_dyn != "":
            attributes_list = self.creation.attribute_line(value=article.materiaux_dyn, number="96")
            qs.appendRow(attributes_list)

        if article.objet != "":
            attributes_list = self.creation.attribute_line(value=article.objet, number="76")
            qs.appendRow(attributes_list)

        if article.quantite != "":
            attributes_list = self.creation.attribute_line(value=article.quantite, number="215")
            qs.appendRow(attributes_list)

        if article.price != "":
            attributes_list = self.creation.attribute_line(value=article.price, number="203")
            qs.appendRow(attributes_list)

    @staticmethod
    def a___________________bdd_bcm_price__________________():
        pass

    def get_all_prices(self):

        file_name = find_filename(file_path=self.chemin_fichier)

        file_name_new = file_name.lower().replace("vw", "pr")

        folder_path = find_folder_path(file_path=self.chemin_fichier)

        file_path = f"{folder_path}{file_name_new}.dbf"

        if not os.path.exists(file_path):
            return

        tps = time.perf_counter()

        try:
            with dbf.Table(filename=file_path) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    price_float = convertir_bytes(record.PREPB)

                    if not isinstance(price_float, float):
                        continue

                    if price_float == 0:
                        continue

                    price_str = f"{price_float:.2f}"

                    current_code = convertir_bytes(record.PRIDPOS)

                    if not isinstance(current_code, str):
                        continue

                    if current_code in self.prices_dict:
                        continue

                    self.prices_dict[current_code] = price_str

        except Exception as error:
            print(f"convert_manage -- ConvertBcmComposants -- get_all_prices -- Erreur : {error}")
            return

        print(f"ConvertBcmComposants -- Price : {time.perf_counter() - tps}s")


class BcmSettings(QWidget):
    bcm_settings_closed = pyqtSignal(bool, bool, bool, bool)

    def __init__(self):
        super().__init__()

        # ---------------------------------------
        # LOADING UI
        # ---------------------------------------

        self.ui = Ui_BcmSettings()
        self.ui.setupUi(self)

        # -----------------------------------------------
        # Settings
        # -----------------------------------------------

        bcm_datas = settings_read(bcm_setting_file)

        if not isinstance(bcm_datas, dict):
            bcm_datas = dict(bcm_setting_datas)

        # -----------------------------------------------

        bcm_material_group = bcm_datas.get("bcm_material_group", bcm_setting_datas.get("bcm_material_group"))

        if not isinstance(bcm_material_group, bool):
            bcm_material_group = bcm_setting_datas.get("bcm_material_group")

        self.ui.bcm_material_group.setChecked(bcm_material_group)

        # -----------------------------------------------

        bcm_link = bcm_datas.get("bcm_link", bcm_setting_datas.get("bcm_link"))

        if not isinstance(bcm_link, bool):
            bcm_link = bcm_setting_datas.get("bcm_link")

        self.ui.bcm_link.setChecked(bcm_link)

        # -----------------------------------------------

        bcm_comment = bcm_datas.get("bcm_comment", bcm_setting_datas.get("bcm_comment"))

        if not isinstance(bcm_comment, bool):
            bcm_comment = bcm_setting_datas.get("bcm_comment")

        self.ui.bcm_comment.setChecked(bcm_comment)

        # -----------------------------------------------

        bcm_price = bcm_datas.get("bcm_price", bcm_setting_datas.get("bcm_price"))

        if not isinstance(bcm_price, bool):
            bcm_price = bcm_setting_datas.get("bcm_price")

        self.ui.bcm_price.setChecked(bcm_price)

        # -----------------------------------------------

        self.ui.ok.clicked.connect(self.bcm_saved)
        self.ui.quit.clicked.connect(self.close)

    def bcm_show(self):

        self.show()

        self.ui.ok.setFocus()

    def bcm_save_settings(self):

        bcm_datas = settings_read(bcm_setting_file)

        if not isinstance(bcm_datas, dict):
            bcm_datas = dict(bcm_setting_datas)

        bcm_datas["bcm_material_group"] = self.ui.bcm_material_group.isChecked()
        bcm_datas["bcm_link"] = self.ui.bcm_link.isChecked()
        bcm_datas["bcm_comment"] = self.ui.bcm_comment.isChecked()
        bcm_datas["bcm_price"] = self.ui.bcm_price.isChecked()

        settings_save(file_name=bcm_setting_file, config_datas=bcm_datas)

    def bcm_saved(self):

        self.bcm_settings_closed.emit(self.ui.bcm_material_group.isChecked(),
                                      self.ui.bcm_link.isChecked(),
                                      self.ui.bcm_comment.isChecked(),
                                      self.ui.bcm_price.isChecked())

    @staticmethod
    def a___________________event______():
        pass

    def keyPressEvent(self, event: QKeyEvent):

        super().keyPressEvent(event)

        if event.key() == Qt.Key_Escape:
            self.close()
            return

    def closeEvent(self, event: QCloseEvent):

        self.bcm_save_settings()

        super().closeEvent(event)


def a___________________allmetre______():
    pass


class AllmetreArticle:
    id_ele = ""
    type_ele = ""
    code = ""
    description = ""
    txt_long = ""
    unite = ""
    formule = ""
    materiaux_dyn = ""
    objet = ""
    quantite = ""

    index_parent = ""
    son_index = ""

    liste_enfants = list()
    liste_enfants_type = list()

    son_parent = None
    son_parent_type = ""

    creation_dossier = False


class ConvertAllmetre(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)
    errors_signal = pyqtSignal(list)

    def __init__(self, allplan, file_path: str, bdd_title: str, conversion=False):
        super().__init__()

        self.allplan = allplan

        if conversion:
            self.allplan.creation.attributes_datas.clear()

        self.creation = self.allplan.creation

        self.chemin_fichier = file_path

        self.conversion = conversion

        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.liste_niveau_zero = list()

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

    def run(self):

        if not self.allmetre_chargement_bdd():
            self.loading_completed.emit(self.model, list(), list())

            return

        tps = time.perf_counter()

        self.allmetre_treeview_creation()

        print(f"ConvertAllmetre : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())

    @staticmethod
    def a___________________bdd_allmetre_creation__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def allmetre_chargement_bdd(self):

        dict_table = dict()
        code_depart = ""
        code_niveau_zero = ""
        inconnu_index = 1

        try:
            with dbf.Table(filename=self.chemin_fichier) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    code_type_ele: str = convertir_bytes(record.VWTYP)

                    if code_type_ele == "X":
                        code_niveau_zero = convertir_bytes(record.VWCID)
                        continue

                    if code_type_ele != "E" and code_type_ele != "P":
                        continue

                    if code_type_ele == "P":
                        type_ele = component_code
                    else:
                        type_ele = folder_code

                    # Création de l'article
                    nouvel_article = AllmetreArticle()
                    nouvel_article.id_ele = convertir_bytes(record.VWSRT)

                    nouvel_article.type_ele = type_ele

                    code = convertir_bytes(record.VWCTX)

                    if code == "":
                        a = self.tr("code vide")
                        code = f'{a} {inconnu_index}'
                        inconnu_index += 1

                    nouvel_article.code = code
                    nouvel_article.description = convertir_bytes(record.VWKTX)
                    nouvel_article.txt_long = convertir_bytes(record.VWLTX)

                    nouvel_article.unite = self.allplan.convert_unit(convertir_bytes(record.VWDIM))

                    formule_tps = convertir_bytes(record.VWATR)

                    formule_metre, materiaux_dyn, objet, quantite = self.allplan.convertir_formule_bdd(formule_tps)

                    nouvel_article.formule = formule_metre
                    nouvel_article.materiaux_dyn = materiaux_dyn
                    nouvel_article.objet = objet
                    nouvel_article.quantite = quantite

                    nouvel_article.liste_enfants = list()
                    nouvel_article.liste_enfants_type = list()
                    nouvel_article.creation_dossier = False

                    index_parent = convertir_bytes(record.VWPID)

                    son_index = convertir_bytes(record.VWCID)

                    if index_parent == "":
                        index_parent = code_niveau_zero

                    # Si niveau 1 -> article à ajouter dans la liste des dossiers à créer dans invisiblerootitem
                    if index_parent == code_niveau_zero:
                        self.liste_niveau_zero.append(nouvel_article)

                        if len(dict_table) == 0 and type_ele == folder_code:
                            code_depart = son_index

                    # Vérification que l'index n'est pas dans le dict --> erreur sinon
                    if son_index in dict_table:
                        print(f"ConvertirBddAllmetre -- allmetre_chargement_bdd -- "
                              f"index {son_index} est déjà dans le dict")
                        continue

                    # Ajout de l'article dans le dict
                    dict_table[son_index] = nouvel_article

                    # Si article du niveau 1 --> passer la recherche de parent
                    if index_parent == code_niveau_zero:
                        continue

                    # Vérification que l'index du parent existe dans le dict --> erreur sinon
                    if index_parent not in dict_table:
                        print(f"ConvertirBddAllmetre -- allmetre_chargement_bdd -- "
                              f"index parent {index_parent} n'est pas dans le dict")
                        continue

                    # Mise en mémoire du parent
                    article_parent = dict_table[index_parent]

                    if type_ele == component_code and article_parent.type_ele != material_code:
                        article_parent.type_ele = material_code

                    # Ajout de l'article dans la liste des enfants du parent actuel
                    liste_enfants = article_parent.liste_enfants
                    liste_enfants.append(nouvel_article)

                    # Ajout du parent et du type de parent dans l'article actuel
                    nouvel_article.son_parent = article_parent
                    nouvel_article.son_parent_type = article_parent.type_ele

        except Exception as erreur:

            self.loading_completed.emit(self.model, list(), list())
            print(f"widget_onglet -- QThreadBddAllmetre -- allmetre_chargement_bdd -- {erreur}")
            return False

        if code_depart == "":
            self.loading_completed.emit(self.model, list(), list())

            print(f"widget_onglet -- QThreadBddAllmetre -- allmetre_chargement_bdd -- code de départ est vide")
            return False

        if len(dict_table) == 0:
            self.loading_completed.emit(self.model, list(), list())

            print(f"widget_onglet -- QThreadBddAllmetre -- allmetre_chargement_bdd -- dict_table est vide")
            return False

        return True

    def allmetre_treeview_creation(self):

        hierarchie_parent = self.model.invisibleRootItem()

        for article_actuel in self.liste_niveau_zero:

            article_actuel: AllmetreArticle

            # 001 ajouter dossier (mystandarditem) dans hierarchie_parent
            mystandarditem: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                              article_actuel.code,
                                                              article_actuel.description)

            self.allmetre_treeview_analyser_enfants(article_actuel)

            if article_actuel.type_ele == material_code:
                mystandarditem_dossier = mystandarditem

                # 002 ajouter ouvrage (mystandarditem) dans mystandarditem_dossier
                mystandarditem: MyQstandardItem = self.material_add(mystandarditem_dossier,
                                                                    article_actuel)

            self.allmetre_treeview_charger_enfants(article_actuel, mystandarditem)

    def allmetre_treeview_analyser_enfants(self, actuel_article: AllmetreArticle):

        actuel_liste_enfants = actuel_article.liste_enfants

        if len(actuel_liste_enfants) == 0:
            return

        # actuel_description = actuel_article.description
        actuel_type_ele = actuel_article.type_ele

        if actuel_type_ele == component_code:
            return

        liste_ele_enfants_type = list()

        # print(f"############## {actuel_description} -- {actuel_type_ele} ############## ")

        for analyse_enfant in actuel_liste_enfants:
            analyse_type_ele = analyse_enfant.type_ele

            liste_ele_enfants_type.append(analyse_type_ele)
            self.allmetre_treeview_analyser_enfants(analyse_enfant)

        if component_code in liste_ele_enfants_type and \
                (material_code in liste_ele_enfants_type or folder_code in liste_ele_enfants_type):

            indices = [i for i, x in enumerate(liste_ele_enfants_type) if x == component_code]

            for index_a_modifier in reversed(indices):

                try:
                    actuel_article.liste_enfants.pop(index_a_modifier)
                    liste_ele_enfants_type.pop(index_a_modifier)
                    actuel_article.type_ele = folder_code
                except IndexError as erreur:
                    print(erreur)
                    pass

        if material_code in liste_ele_enfants_type and folder_code in liste_ele_enfants_type:

            indices = [i for i, x in enumerate(liste_ele_enfants_type) if x == folder_code]

            verification_si_tous_dossiers_vides = True

            for index_a_verifier in indices:

                article_a_verifier: AllmetreArticle = actuel_liste_enfants[index_a_verifier]

                if len(article_a_verifier.liste_enfants) > 0:
                    verification_si_tous_dossiers_vides = False
                    break

            if not verification_si_tous_dossiers_vides:

                indices = [i for i, x in enumerate(liste_ele_enfants_type) if x == material_code]

                for index_a_modifier in indices:
                    article_a_modifier: AllmetreArticle = actuel_liste_enfants[index_a_modifier]
                    article_a_modifier.creation_dossier = True

                # print(f"Dans {actuel_article.description} -- Création d'un dossier  pour {len(indices)} ouvrage(s)")

            else:

                for index_a_modifier in indices:
                    article_a_modifier: AllmetreArticle = actuel_liste_enfants[index_a_modifier]
                    article_a_modifier.type_ele = material_code

        actuel_article.liste_enfants_type = liste_ele_enfants_type

        # print(f"{actuel_article.description} -- {actuel_type_ele} -- {liste_ele_enfants_type}")
        # print(f"***************************************************************************************************")

    def allmetre_treeview_charger_enfants(self, actuel_article: AllmetreArticle, hierarchie_parent: MyQstandardItem):

        actuel_liste_enfants = actuel_article.liste_enfants

        if len(actuel_liste_enfants) == 0:
            return

        for analyse_enfant in actuel_liste_enfants:

            analyse_enfant: AllmetreArticle

            analyse_code = analyse_enfant.code
            analyse_description = analyse_enfant.description
            analyse_type_ele = analyse_enfant.type_ele
            analyse_creation_dossier = analyse_enfant.creation_dossier

            if analyse_type_ele == folder_code:

                if not isinstance(hierarchie_parent, Folder):

                    print(f"AllmetreConvertirBdd -- allmetre_treeview_charger_enfants -- "
                          f"erreur ajout {analyse_type_ele} -- {analyse_enfant.description} "
                          f"dans dossier : {hierarchie_parent.text()}")
                    continue

                else:

                    # 003 ajouter dossier (mystandarditem) dans hierarchie_parent
                    mystandarditem: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                      analyse_code,
                                                                      analyse_description)

            elif analyse_type_ele == material_code:

                if analyse_creation_dossier:

                    # 004 ajouter dossier (mystandarditem) dans hierarchie_parent
                    mystandarditem: MyQstandardItem = self.folder_add(hierarchie_parent,
                                                                      analyse_code,
                                                                      analyse_description)

                    # 005 ajouter ouvrage (mystandarditem_ouvrage) dans mystandarditem_dossier
                    mystandarditem_ouvrage: MyQstandardItem = self.material_add(mystandarditem,
                                                                                analyse_enfant)

                    self.allmetre_treeview_charger_enfants(analyse_enfant, mystandarditem_ouvrage)

                else:

                    if not isinstance(hierarchie_parent, Folder):

                        print(f"AllmetreConvertirBdd -- allmetre_treeview_charger_enfants -- "
                              f"erreur ajout {analyse_type_ele} -- {analyse_enfant.description} "
                              f"dans dossier : {hierarchie_parent.text()}")
                        continue

                    else:

                        # 006 ajouter ouvrage (mystandarditem) dans hierarchie_parent
                        mystandarditem: MyQstandardItem = self.material_add(hierarchie_parent,
                                                                            analyse_enfant)

            elif analyse_type_ele == component_code:

                if not isinstance(hierarchie_parent, Material):

                    print(f"AllmetreConvertirBdd -- allmetre_treeview_charger_enfants -- "
                          f"erreur ajout {analyse_type_ele} -- {analyse_enfant.description} "
                          f"dans ouvrage : {hierarchie_parent.text()}")

                    continue

                else:

                    # 007 ajouter composant (mystandarditem) dans hierarchie_parent
                    mystandarditem: MyQstandardItem = self.component_add(hierarchie_parent,
                                                                         analyse_enfant)

            else:
                print("AllmetreConvertirBdd -- allmetre_treeview_charger_enfants -- erreur")
                continue

            if not analyse_creation_dossier:
                self.allmetre_treeview_charger_enfants(analyse_enfant, mystandarditem)

            # print(analyse_description)

    @staticmethod
    def a___________________bdd_allmetre_objets__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def folder_add(self, his_parent: MyQstandardItem, code: str, description: str) -> MyQstandardItem:

        # print(f"creation_dossier = {code} - {description}")

        if code.startswith(self.tr("code vide")):
            qs_list = self.creation.folder_line(value=description,
                                                tooltips=False)
        else:
            qs_list = self.creation.folder_line(value=code,
                                                description=description,
                                                tooltips=False)

        his_parent.appendRow(qs_list)
        return qs_list[col_cat_value]

    def material_add(self, his_parent: MyQstandardItem, article: AllmetreArticle) -> MyQstandardItem:

        qs_list = self.creation.material_line(value=article.code,
                                              description=article.description,
                                              tooltips=False)

        qs: MyQstandardItem = qs_list[col_cat_value]

        self.attributes_add(qs=qs,
                            article=article)

        his_parent.appendRow(qs_list)
        return qs

    def component_add(self, his_parent: MyQstandardItem, article: AllmetreArticle) -> MyQstandardItem:

        qs_list = self.creation.component_line(value=f"{article.code}",
                                               description=article.description,
                                               tooltips=False)

        qs: MyQstandardItem = qs_list[col_cat_value]

        self.attributes_add(qs=qs, article=article)

        his_parent.appendRow(qs_list)
        return qs

    def attributes_add(self, qs: MyQstandardItem, article: AllmetreArticle):

        if article.txt_long != "":
            attributes_list = self.creation.attribute_line(value=article.txt_long, number="208")
            qs.appendRow(attributes_list)

        if article.unite != "":
            attributes_list = self.creation.attribute_line(value=article.unite, number="202")
            qs.appendRow(attributes_list)

        if article.formule != "":
            attributes_list = self.creation.attribute_line(value=article.formule, number="267")
            qs.appendRow(attributes_list)

        if article.materiaux_dyn != "":
            attributes_list = self.creation.attribute_line(value=article.materiaux_dyn, number="96")
            qs.appendRow(attributes_list)

        if article.objet != "":
            attributes_list = self.creation.attribute_line(value=article.objet, number="76")
            qs.appendRow(attributes_list)

        if article.quantite != "":
            attributes_list = self.creation.attribute_line(value=article.quantite, number="215")
            qs.appendRow(attributes_list)


@dataclass
class Article:
    libel: str
    unit: str
    codecm: str
    codest: str
    odr: str
    code_art: str


@dataclass
class SousDossier:
    codest: str
    libel: str
    odr: str
    articles: List[Article] = field(default_factory=list)


@dataclass
class Dossier:
    codecm: str
    libel: str
    odr: str
    sous_dossiers: Dict[str, SousDossier] = field(default_factory=dict)


def a___________________gimi______():
    pass


class ConvertGimi(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)
    errors_signal = pyqtSignal(str, str, str)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan = allplan

        self.creation = self.allplan.creation

        file_path = os.path.dirname(file_path)

        file_path = file_path.replace("/", '\\')

        if not file_path.endswith("\\"):
            file_path += "\\"

        self.chemin_fichier = file_path

        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.titre_message = "GIMI"

    def run(self):

        tps = time.perf_counter()

        self.chargement()

        print(f"ConvertGimi : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())

    def chargement(self) -> bool:

        file_txt = self.tr("Le fichier")
        find_txt = self.tr("n'a pas été trouvé")

        if not os.path.exists(self.chemin_fichier):
            self.errors_signal.emit(self.titre_message,
                                    "Le chemin d'allmétré n'a pas été trouvé",
                                    self.chemin_fichier)

            return False

        chemin_bdd_cm_fic = f"{self.chemin_fichier}CM.FIC"

        if not os.path.exists(chemin_bdd_cm_fic):
            print(f"Le fichier chemin_bdd_cm_fic : {chemin_bdd_cm_fic} n'a pas été trouvé, "
                  f"la conversion de la BDD ne peut se faire, désolé")

            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} cm.fic {find_txt}",
                                    chemin_bdd_cm_fic)

            return False

        chemin_bdd_cm_ndx = f"{self.chemin_fichier}CM.NDX"

        if not os.path.exists(chemin_bdd_cm_ndx):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} cm.ndx {find_txt}",
                                    chemin_bdd_cm_ndx)
            return False

        chemin_bdd_st_fic = f"{self.chemin_fichier}ST.FIC"

        if not os.path.exists(chemin_bdd_st_fic):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} st.fic {find_txt}",
                                    chemin_bdd_st_fic)
            return False

        chemin_bdd_st_ndx = f"{self.chemin_fichier}ST.NDX"

        if not os.path.exists(chemin_bdd_st_ndx):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} st.ndx {find_txt}",
                                    chemin_bdd_st_ndx)
            return False

        chemin_bdd_articles_fic = f"{self.chemin_fichier}ARTICLES.FIC"

        if not os.path.exists(chemin_bdd_articles_fic):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} articles.fic {find_txt}",
                                    chemin_bdd_articles_fic)
            return False

        chemin_bdd_articles_ndx = f"{self.chemin_fichier}ARTICLES.NDX"

        if not os.path.exists(chemin_bdd_articles_ndx):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} articles.ndx {find_txt}",
                                    chemin_bdd_articles_ndx)

            return False

        chemin_bdd_articles_mmo = f"{self.chemin_fichier}ARTICLES.MMO"

        if not os.path.exists(chemin_bdd_articles_mmo):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} articles.mmo {find_txt}",
                                    chemin_bdd_articles_mmo)

            return False

        chemin_export_bdd_cm = f"{asc_export_path}CM.xml"
        chemin_export_bdd_st = f"{asc_export_path}ST.xml"
        chemin_export_bdd_articles = f"{asc_export_path}ARTICLES.xml"

        chemin_hyperfile2xml = f"{asc_exe_path}tools\\hyperfile2xml.exe"

        if not os.path.exists(chemin_hyperfile2xml):
            self.errors_signal.emit(self.titre_message,
                                    f"{file_txt} hyperfile2xml.exe {find_txt}",
                                    chemin_hyperfile2xml)
            return False

        cmd = f'"{chemin_hyperfile2xml}" "{chemin_bdd_cm_fic}" "{chemin_export_bdd_cm}"\n'
        cmd += f'"{chemin_hyperfile2xml}" "{chemin_bdd_st_fic}" "{chemin_export_bdd_st}"\n'
        cmd += f'"{chemin_hyperfile2xml}" "{chemin_bdd_articles_fic}" "{chemin_export_bdd_articles}"\n'

        cmd_path = f"{asc_export_path}convert.bat"

        try:

            if os.path.exists(cmd_path):
                os.remove(cmd_path)

            if os.path.exists(chemin_export_bdd_cm):
                os.remove(chemin_export_bdd_cm)

            if os.path.exists(chemin_export_bdd_st):
                os.remove(chemin_export_bdd_st)

            if os.path.exists(chemin_export_bdd_articles):
                os.remove(chemin_export_bdd_articles)

        except OSError as erreur:

            supp_txt = self.tr("La suppression du fichier convert.bat a échouée")

            self.errors_signal.emit(self.titre_message,
                                    f"{supp_txt}\n",
                                    f"{erreur}")

            return False

        try:
            with open(cmd_path, "w") as file:

                file.writelines(cmd)

        except OSError as erreur:

            write_txt = self.tr("L'écriture du fichier convert.bat a échouée")

            self.errors_signal.emit(self.titre_message,
                                    f"{write_txt}\n",
                                    f"{erreur}")

            return False

        err_txt = self.tr("Une erreur est survenue.")

        try:
            subprocess.call([cmd_path], stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)

        except subprocess.CalledProcessError as erreur:

            self.errors_signal.emit(self.titre_message,
                                    f"{err_txt}\n",
                                    f"{erreur}")

            return False

        except Exception as erreur:

            self.errors_signal.emit(self.titre_message,
                                    f"{err_txt}\n",
                                    f"{erreur}")
            return False

        if not os.path.exists(chemin_export_bdd_cm):
            print(f"Le fichier chemin_export_bdd_cm : {chemin_export_bdd_cm} n'existe pas après la conversion")

            b = self.tr("n'existe pas après la conversion")

            self.errors_signal.emit(self.titre_message,
                                    f"{err_txt}\n",
                                    f"{file_txt} chemin_export_bdd_cm : {chemin_export_bdd_cm} {b}")
            return False

        # -----------------------------------
        # --- LECTURE XML
        # -----------------------------------

        dossiers: Dict[str, Dossier] = {}

        cm_tree = validation_fichier_xml(chemin_export_bdd_cm)

        if cm_tree is None:
            return False

        for element_child_data in cm_tree.findall('Data'):

            libel = element_child_data.findtext("LIBEL")
            codecm = element_child_data.findtext("ID_CM")
            odr: str = element_child_data.findtext("ODR")

            if libel is None or codecm is None or odr is None:
                continue

            dossiers[codecm] = Dossier(codecm=codecm, libel=libel, odr=odr.zfill(10))

        # ------------------------------------------------------

        sous_dossiers: Dict[str, SousDossier] = {}
        st_tree = etree.parse(chemin_export_bdd_st)

        for element_child_data in st_tree.findall('Data'):

            libel = element_child_data.findtext("LIBELST")
            codest = element_child_data.findtext("ID_ST")
            odr = element_child_data.findtext("ODR")

            if libel is None or codest is None or odr is None:
                continue

            sous_dossiers[codest] = SousDossier(codest=codest, libel=libel, odr=odr.zfill(10))

        # ------------------------------------------------------

        articles_tree = etree.parse(chemin_export_bdd_articles)
        for element_child_data in articles_tree.findall('Data'):

            libel = element_child_data.findtext("LIBEL")
            odr = element_child_data.findtext("ODR")
            code_art = element_child_data.findtext("CODE_ART")
            codecm = element_child_data.findtext("CODECM")
            codest = element_child_data.findtext("CODEST")
            unit = element_child_data.findtext("UNIT")

            # On fait le contrôle d'erreur immédiatement pour sauterles articles avec des données incomplètes
            if libel is None \
                    or odr is None \
                    or code_art is None \
                    or codecm is None \
                    or codest is None \
                    or unit is None:
                continue

            if codecm not in dossiers:
                continue

            if codest not in sous_dossiers:
                continue

            dossier = dossiers[codecm]
            sous_dossier = sous_dossiers[codest]

            copie_sous_dossier = dossier.sous_dossiers.get(codest)

            if copie_sous_dossier is None:
                # On duplique note sous-dossier
                copie_sous_dossier = dataclasses.replace(sous_dossier, articles=list())
                # ...et on l'ajoute sous le dossier parent
                dossier.sous_dossiers[codest] = copie_sous_dossier

            article = Article(libel=libel, unit=unit, codecm=codecm, codest=codest, odr=odr.zfill(10),
                              code_art=code_art)

            copie_sous_dossier.articles.append(article)

        # -----------------------------------
        # --- LECTURE MEMOIRE
        # -----------------------------------

        root_model_view = self.model.invisibleRootItem()

        nb_ele = len(dossiers)

        if nb_ele == 0:
            return False

        id_art = 0

        for dossier in dossiers.values():

            id_art += 1

            if not dossier.sous_dossiers:
                continue

            dossier_item_libel: MyQstandardItem = self.folder_add(root_model_view,
                                                                  dossier.libel)

            for sous_dossier in dossier.sous_dossiers.values():

                sous_dossier: SousDossier

                if not sous_dossier.articles:
                    continue

                ss_dossier_item_libel: MyQstandardItem = self.folder_add(dossier_item_libel,
                                                                         sous_dossier.libel)

                for article in sous_dossier.articles:
                    self.component_add(ss_dossier_item_libel, article)

        return True

    @staticmethod
    def a___________________bdd_gimi_objets__________________():
        """ Partie réservée à la gestion des boutons des attributs"""
        pass

    def folder_add(self, his_parent: MyQstandardItem, code: str) -> MyQstandardItem:

        qs_list = self.creation.folder_line(value=code, tooltips=False)
        his_parent.appendRow(qs_list)
        return qs_list[col_cat_value]

    def component_add(self, son_parent: MyQstandardItem, article: Article) -> MyQstandardItem:

        qs_list = self.creation.component_line(value=article.code_art,
                                               description=article.libel,
                                               tooltips=False)

        qs: MyQstandardItem = qs_list[col_cat_value]

        if article.unit != "":
            liste_attribut = self.creation.attribute_line(value=article.unit, number="202")
            qs.appendRow(liste_attribut)

        son_parent.appendRow(qs_list)
        return qs


def a___________________favorites______():
    pass


class ConvertFavorite(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan = allplan

        self.creation = self.allplan.creation
        self.chemin_fichier = file_path
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

    def run(self):

        tps = time.perf_counter()

        try:
            # encodage = recherche_encod(self.chemin_fichier)

            parser = etree.XMLParser(recover=True)
            tree = etree.parse(self.chemin_fichier, parser=parser)
            root = tree.getroot()

            titre = self.tr("Favoris Allplan")

            dict_favoris_allplan = get_favorites_allplan_dict()

            for extension, nom_favoris in dict_favoris_allplan.items():

                if self.chemin_fichier.endswith(extension):
                    titre = f"Favoris Allplan {nom_favoris}"

            list_refus = ['4', '10', '94',
                          '101', '102', '103', '104', '105', '106', '108', '109', '112', '113', '114', '115',
                          '116', '117', '119', '133', '155', '156', '157', '161', '169', '171', '177',
                          '210', '253', '254', '255', '267',
                          '331', '333', '334', '336', '337', '338', '348', '373', '379', '383',
                          "689",
                          '748', '749',
                          '979',
                          '1414', '1442',
                          '1958', '1959', "1960", "1978", "1979",
                          '2100', '2101', '2102', '2103', '2105', '2105', '2110', '2111', '2112', '2113', '2114',
                          '2160',
                          '4302', '4303']

            liste_datas = list()

            qs_list = self.creation.folder_line(value=titre, tooltips=False)
            self.model.invisibleRootItem().appendRow(qs_list)
            folder = qs_list[col_cat_value]

            for attrib_set in root.iter("NEM_ATTRIB_SET"):

                dict_attributs = dict()

                for elem in attrib_set:

                    ifnr = elem.find("IFNR")

                    if not get_value_is_valid(ifnr):
                        continue

                    ifnr_text = ifnr.text

                    if ifnr_text == "":
                        continue

                    if ifnr_text in list_refus:
                        continue

                    value = elem.find("VALUE")

                    if value is None:
                        continue

                    value_text = value.text

                    if value_text is None:
                        value_text = ""

                    dict_attributs[ifnr_text] = value_text

                liste_datas.append(dict_attributs)

        except Exception as erreur:
            msg(titre=application_title,
                message=self.tr("Une erreur est survenue."),
                type_bouton=QMessageBox.Ok,
                icone_critique=True,
                details=f"{erreur}")

            return

        for dict_attributs in liste_datas:

            dict_attributs: dict

            # -----------------------------------
            # Code
            # -----------------------------------

            if attribute_default_base in dict_attributs:
                code: str = dict_attributs[attribute_default_base]

                if code.strip() == "":

                    if "508" in dict_attributs:
                        code: str = dict_attributs["508"]
                    else:
                        continue
            else:
                continue

            code = code.strip()

            if code == "":
                continue

            # -----------------------------------
            # Description
            # -----------------------------------

            description = dict_attributs.get("207", "")

            # -----------------------------------
            # Material Création
            # -----------------------------------

            liste_ouvrage = self.creation.material_line(value=code, description=description, tooltips=False)

            folder.appendRow(liste_ouvrage)

            ouvrage: MyQstandardItem = liste_ouvrage[col_cat_value]

            liste_actuelle = list()

            # -----------------------------------
            # Attributes creation
            # -----------------------------------

            for number, value in dict_attributs.items():

                if number in liste_actuelle:
                    continue

                # -----------------------------------
                # Layer
                # -----------------------------------

                if number == attribute_val_default_layer_first:

                    for layer_number in attribute_val_default_layer:

                        if layer_number in liste_actuelle:
                            continue

                        if layer_number in dict_attributs:

                            value = dict_attributs.get(layer_number, "")

                        else:

                            value = attribute_val_default_layer.get(layer_number, "")

                        liste_actuelle.append(layer_number)

                        liste_attribut = self.creation.attribute_line(value=value,
                                                                      number=layer_number,
                                                                      model_enumeration=None)

                        index_insert = ouvrage.get_attribute_insertion_index(number=layer_number)
                        ouvrage.insertRow(index_insert, liste_attribut)

                    continue

                # -----------------------------------
                #  Fill
                # -----------------------------------

                if number == attribute_val_default_fill_first:

                    for fill_number in attribute_val_default_fill:

                        if fill_number in liste_actuelle:
                            continue

                        if fill_number in dict_attributs:

                            value = dict_attributs.get(fill_number, "")

                        else:

                            value = attribute_val_default_fill.get(fill_number, "")

                        if fill_number == "111":

                            id_hachurage = dict_attributs.get("118", attribute_val_default_fill.get(number, "0"))

                            if id_hachurage == "1":
                                model_c = self.allplan.model_haching

                            elif id_hachurage == "2":
                                model_c = self.allplan.model_pattern

                            elif id_hachurage == "3":
                                model_c = self.allplan.model_color

                            elif id_hachurage == "5":
                                model_c = self.allplan.model_style

                            else:
                                model_c = None

                        else:

                            model_c = None

                        liste_actuelle.append(fill_number)

                        liste_attribut = self.creation.attribute_line(value=value,
                                                                      number=fill_number,
                                                                      model_enumeration=model_c)

                        index_insert = ouvrage.get_attribute_insertion_index(number=fill_number)
                        ouvrage.insertRow(index_insert, liste_attribut)

                    continue

                # -----------------------------------
                #  ROOM
                # -----------------------------------

                if number == attribute_val_default_room_first:

                    for room_number in attribute_val_default_room:

                        if room_number in liste_actuelle:
                            continue

                        if room_number in dict_attributs:

                            value = dict_attributs.get(room_number, "")

                        else:

                            value = attribute_val_default_room.get(room_number, "")

                        liste_actuelle.append(room_number)

                        liste_attribut = self.creation.attribute_line(value=value,
                                                                      number=room_number,
                                                                      model_enumeration=None)

                        index_insert = ouvrage.get_attribute_insertion_index(number=room_number)
                        ouvrage.insertRow(index_insert, liste_attribut)

                    continue

                # -----------------------------------
                #  OTHER
                # -----------------------------------

                liste_actuelle.append(number)

                liste_attribut = self.creation.attribute_line(value=value,
                                                              number=number,
                                                              model_enumeration=None)

                index_insert = ouvrage.get_attribute_insertion_index(number=number)
                ouvrage.insertRow(index_insert, liste_attribut)

        print(f"ConvertFavorite : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())


def a___________________kukat______():
    pass


class ConvertKukat(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)
    errors_signal = pyqtSignal(list)

    def __init__(self, allplan, file_path: str, bdd_title: str, conversion=False):
        super().__init__()

        self.allplan = allplan

        self.creation = self.allplan.creation

        self.dossier = find_folder_path(file_path)

        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.chapitre = "chapters"
        self.extension = ".KAT"

        self.liste_code = list()
        self.datas = dict()

        self.conversion = conversion

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

    def run(self):

        tps = time.perf_counter()

        self.datas = self.lecture_fichier_cat(f"{self.dossier}{self.chapitre}{self.extension}")

        self.lecture_enfants(self.datas)

        self.convert_to_model(self.datas, self.model.invisibleRootItem())

        print(f"ConvertKukat : {time.perf_counter() - tps}s")

        self.loading_completed.emit(self.model, list(), list())

    @staticmethod
    def lecture_fichier_cat(file_path: str, ouvrage=False) -> dict:

        if not os.path.exists(file_path):
            return dict()

        lines_list = read_file_to_list(file_path=file_path)

        resultats = dict()

        if ouvrage:
            max_split = 2
        else:
            max_split = 1

        try:
            for ligne in lines_list[3:]:
                partie = ligne.strip().split(None, max_split)
                if len(partie) == 2 and not ouvrage:
                    reference, description = partie
                    resultats[reference] = {"description": description, "enfants": {}}

                elif len(partie) == 3 and ouvrage:
                    reference, unite, description = partie
                    resultats[reference] = {"code": reference, "description": description, "unite": unite, }

            return resultats

        except Exception as erreur:
            print(f"lecture_fichier_kat erreur -- décompose: {erreur}")
            return dict()

    def lecture_enfants(self, data: dict):

        for key in data:

            fichier_actuel = f"{self.dossier}{key}{self.extension}"

            if not os.path.exists(fichier_actuel):
                continue

            if "enfants" not in data[key]:
                return

            if len(key) == 3:
                data[key]["ouvrage"] = self.lecture_fichier_cat(fichier_actuel, ouvrage=True)

            else:
                data[key]["enfants"] = self.lecture_fichier_cat(fichier_actuel)
                self.lecture_enfants(data[key]["enfants"])

    def convert_to_model(self, data: dict, qs_dossier_parent: QStandardItem):

        for key, enfants in data.items():

            enfants: dict

            description = enfants.get("description", "")

            # -----------
            # Dossier
            # -----------

            if "enfants" in enfants:

                liste_qs_dossier_enfant = self.creation.folder_line(value=key,
                                                                    description=description,
                                                                    tooltips=self.conversion)

                qs_dossier_enfant = liste_qs_dossier_enfant[0]

                data_enfants = enfants["enfants"]

                # -----------
                # Ouvrage
                # -----------

                if "ouvrage" in enfants:
                    self.material_add(enfants["ouvrage"], qs_dossier_enfant)

                if len(data_enfants) != 0:
                    self.convert_to_model(data_enfants, qs_dossier_enfant)

                qs_dossier_parent.appendRow(liste_qs_dossier_enfant)

    def material_add(self, data: dict, qs_dossier_parent: MyQstandardItem):

        for key, enfants in data.items():

            enfants: dict

            if key in self.liste_code:
                return

            description = enfants.get("description", "")
            unite = enfants.get("unite", "")

            liste_qs_ouvrage = self.creation.material_line(value=key, description=description)

            qs_ouvrage: MyQstandardItem = liste_qs_ouvrage[0]

            if unite != "":
                liste_qs_attribut = self.creation.attribute_line(value=unite, number="202")
                qs_ouvrage.appendRow(liste_qs_attribut)

            qs_dossier_parent.appendRow(liste_qs_ouvrage)

            self.liste_code.append(key)


def a___________________extern______():
    pass


class ConvertExtern(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan: AllplanDatas = allplan

        self.creation = self.allplan.creation

        self.file_path = file_path

        self.model_cat = QStandardItemModel()
        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])
        self.datas = dict()

        self.root = None

        self.conversion_code = {"description": "207",
                                "unité": "202",
                                "prix": "203",
                                "description_contrat": "501",
                                "option": "502",
                                "variant": "503",
                                "niveau": "504"}

    def run(self):

        tps = time.perf_counter()

        if self.file_path.startswith("http"):
            try:
                with urllib.request.urlopen(self.file_path) as response:
                    xml = response.read().decode()

                    if isinstance(xml, str):
                        self.root = etree.fromstring(xml)

                    else:
                        print(f"widget_bible_externe_onglet -- BddExtern -- erreur chargement : not str")
                        self.loading_completed.emit(self.model_cat, list(), list())
                        return

            except Exception as erreur:
                print(f"widget_bible_externe_onglet -- BddExtern -- erreur chargement : {erreur}")
                self.loading_completed.emit(self.model_cat, list(), list())
                return

        else:

            if not os.path.exists(self.file_path):
                print(f"widget_bible_externe_onglet -- BddExtern -- erreur chargement : not os.path.exists(file_path)")
                self.loading_completed.emit(self.model_cat, list(), list())
                return

            try:
                tree = etree.parse(self.file_path)
                self.root = tree.getroot()

            except Exception as error:
                print(f"catalog_manage -- CatalogLoad --  analyse_display -- {error}")

                self.loading_completed.emit(self.model_cat, list(), list())

                return False

        self.model_cat.invisibleRootItem().setData(folder_code, user_data_type)

        # -------------------------------------------------
        # chargement hierarchie
        # -------------------------------------------------

        self.model_cat.beginResetModel()

        search_version = self.root.find("Dossier")

        try:
            if search_version is not None:
                self.catalog_load_old(self.model_cat.invisibleRootItem(), self.root)
            else:
                self.catalog_load(self.model_cat.invisibleRootItem(), self.root)

        except Exception as error:
            print(f"catalog_manage -- CatalogLoad --  analyse_display -- {error}")

            self.model_cat.endResetModel()

            self.loading_completed.emit(self.model_cat, list(), list())
            return

        self.model_cat.endResetModel()

        self.loading_completed.emit(self.model_cat, list(), list())

        print(f"ConvertExtern : {time.perf_counter() - tps}s")

    def catalog_load(self, qs_parent: QStandardItem, element: etree._Element):

        for child in element:

            tag = child.tag

            if not isinstance(tag, str):
                print("convert_manage -- ConvertExtern -- catalog_load -- not isinstance(tag, str)")
                continue

            tag = tag.capitalize()

            if tag == folder_code or tag == material_code or tag == component_code:

                if not self.verif_possibility(id_parent=id(qs_parent), type_ele=tag):
                    print("convert_manage -- ConvertExtern -- catalog_load -- not self.verif_possibility")
                    continue

            # -----------------------------------------------
            # Node
            # -----------------------------------------------

            if tag == folder_code:

                name = child.get('name')

                if name is None:
                    print("convert_manage -- ConvertExtern -- catalog_load -- name is None")
                    continue

                description = child.get('description', "")

                if "\n" in description:
                    description = description.replace("\n", "")

                qs_list = self.creation.folder_line(value=name,
                                                    description=description,
                                                    tooltips=False)

                qs_current = qs_list[0]

                self.catalog_load(qs_current, child)

                qs_parent.appendRow(qs_list)

                continue

            # -----------------------------------------------
            # Group and Position
            # -----------------------------------------------

            if tag == material_code or tag == component_code:
                self.children_load(child, qs_parent, tag)
                continue

    def children_load(self, child, qs_parent: QStandardItem, tag: str):

        presence_layer = False
        presence_remplissage = False
        presence_piece = False
        # presence_ht = False

        liste_defaut = list()
        datas_attribut_layer = dict(attribute_val_default_layer)
        datas_attribut_remp = dict(attribute_val_default_fill)
        datas_attribut_piece = dict(attribute_val_default_room)
        # datas_attribut_ht = dict(attribut_val_defaut_ht)

        liste_autres = list()

        liste_attributs = list()

        name = child.get("name", "")
        description = child.get("description", "")
        unit = child.get("unit", "")

        if unit != "":
            liste_defaut.append(["202", self.allplan.convert_unit(unit=unit)])

        # ----------------------------
        # search attributes
        # ----------------------------

        attributes = child.findall('Attribute')

        if len(attributes) != 0:

            for attribute in attributes:

                number = attribute.get("id")
                value = attribute.get("value", "")

                if number is None:
                    print("convert_manage -- ConvertExtern -- children_load -- number is None")
                    return None

                if number in liste_attributs:
                    print("convert_manage -- ConvertExtern -- children_load -- number in liste_attributs")
                    continue

                liste_attributs.append(number)

                # -----------------------------------------
                # Attribute 83
                # -----------------------------------------

                if number == attribute_default_base or number == "207" or number == "202":
                    continue

                # -----------------------------------------
                # Attribute 120  - 209 - 110
                # -----------------------------------------

                if number in attribut_val_defaut_defaut:
                    liste_defaut.append([number, value])
                    continue

                # -----------------------------------------
                # Attribute 141 - 349 - 346 - 345 - 347
                # -----------------------------------------

                if number in attribute_val_default_layer:
                    presence_layer = True
                    datas_attribut_layer[number] = value
                    continue

                # -----------------------------------------
                # Attribute 118 - 111 - 252 - 336 - 600
                # -----------------------------------------

                if number in attribute_val_default_fill:
                    presence_remplissage = True
                    datas_attribut_remp[number] = value
                    continue

                # -----------------------------------------
                # Attribute 231 - 235 - 232 - 266 - 233 - 264
                # -----------------------------------------

                if number in attribute_val_default_room:
                    presence_piece = True

                    if number == "232":
                        datas_attribut_piece[number] = self.allplan.traduire_valeur_232(value_current=value)

                    elif number == "233":
                        datas_attribut_piece[number] = self.allplan.traduire_valeur_233(value_current=value)

                    elif number == "235":
                        datas_attribut_piece[number] = self.allplan.traduire_valeur_235(value_current=value)

                    else:
                        datas_attribut_piece[number] = value

                    continue

                # -----------------------------------------
                # Attribute 112 - 113 - 114 - 115 - 169 - 171 - 1978 - 1979
                # -----------------------------------------

                # if number in attribut_val_defaut_ht:
                #     presence_ht = True
                #     datas_attribut_ht[number] = value
                #     continue

                # -----------------------------------------
                # Attribute 76 - 96 - 180 - 267
                # -----------------------------------------

                if number in formula_list_attributes:

                    if self.allplan.version_allplan_current != "2022":

                        if len(re.findall(pattern=formula_piece_pattern, string=value)):
                            value = re.sub(pattern=formula_piece_pattern,
                                           repl=lambda m: formula_piece_dict.get(m.group(0)),
                                           string=value,
                                           flags=re.IGNORECASE)

                    liste_autres.append([number, value])
                    continue

                # -----------------------------------------
                # Attribute > 1999 & < 12000
                # -----------------------------------------

                try:
                    number_int = int(number)

                    if 1999 < number_int < 12000:

                        datas_attribute = self.allplan.attributes_dict.get(number, dict())

                        type_attribut = datas_attribute.get(code_attr_option, "")

                        if type_attribut not in [code_attr_formule_str, code_attr_formule_int, code_attr_formule_float]:
                            liste_autres.append([number, value])
                            continue

                        valeur_formule = datas_attribute.get(code_attr_value, "")

                        if self.allplan.version_allplan_current != "2022":
                            if len(re.findall(pattern=formula_piece_pattern, string=valeur_formule)):
                                valeur_formule = re.sub(pattern=formula_piece_pattern,
                                                        repl=lambda m: formula_piece_dict.get(m.group(0)),
                                                        string=valeur_formule,
                                                        flags=re.IGNORECASE)

                        liste_autres.append([number, valeur_formule])
                        continue

                except ValueError as error:
                    print(f"convert_manage -- ConvertExtern -- children_load -- {error}")
                    pass

                # -----------------------------------------
                # Attribute Other
                # -----------------------------------------

                liste_autres.append([number, value])
                continue

            liste_defaut.sort(key=lambda x: int(x[0]))
            liste_autres.sort(key=lambda x: int(x[0]))

        # ==================================================================== #
        # QStandardItem Creation (Material / Component
        # ==================================================================== #

        if name == "":
            print("convert_manage -- ConvertExtern -- children_load -- name is empty")
            return None

        if tag == material_code:

            qs_list = self.creation.material_line(value=name,
                                                  description=description,
                                                  tooltips=False)

            qs_current: QStandardItem = qs_list[0]

        elif tag == component_code:

            qs_list = self.creation.component_line(value=name,
                                                   description=description,
                                                   tooltips=False)

            qs_current: QStandardItem = qs_list[0]

        else:
            print(f"convert_manage -- ConvertExtern -- children_load -- tag is wrong : {tag}")
            return None

        # ==================================================================== #
        # Attribute Creation
        # ==================================================================== #

        # -----------------------------------------
        # Attribute 120  - 209 - 110
        # -----------------------------------------

        for number, value in liste_defaut:
            qs_current.appendRow(self.creation.attribute_line(value=value, number=number))

        # -----------------------------------------
        # Attribute 118 - 111 - 252 - 336 - 600
        # -----------------------------------------

        if presence_remplissage:

            type_remplissage = datas_attribut_remp.get("118", "0")

            if type_remplissage == "1":
                model_enumeration = self.allplan.model_haching
            elif type_remplissage == "2":
                model_enumeration = self.allplan.model_pattern
            elif type_remplissage == "3":
                model_enumeration = self.allplan.model_color
            elif type_remplissage == "5":
                model_enumeration = self.allplan.model_style
            elif type_remplissage == "6":
                model_enumeration = self.allplan.model_none
            else:
                model_enumeration = self.allplan.model_none
                datas_attribut_remp["111"] = "-1"
                datas_attribut_remp["252"] = "-1"
                datas_attribut_remp["336"] = ""
                datas_attribut_remp["600"] = "0"

            for number, value in datas_attribut_remp.items():

                if number == "111":
                    qs_current.appendRow(self.creation.attribute_line(value=value,
                                                                      number=number,
                                                                      model_enumeration=model_enumeration))
                    continue

                qs_current.appendRow(self.creation.attribute_line(value=value, number=number))

        # -----------------------------------------
        # Attribute 141 - 349 - 346 - 345 - 347
        # -----------------------------------------

        if presence_layer:
            for number, value in datas_attribut_layer.items():
                qs_current.appendRow(self.creation.attribute_line(value=value, number=number))

        # -----------------------------------------
        # Attribute 231 - 235 - 232 - 266 - 233 - 264
        # -----------------------------------------

        if presence_piece:

            for number, value in datas_attribut_piece.items():
                qs_current.appendRow(self.creation.attribute_line(value=value, number=number))

        # -----------------------------------------
        # Attributes Other
        # -----------------------------------------

        if len(liste_autres) != 0:
            for number, value in liste_autres:
                qs_current.appendRow(self.creation.attribute_line(value=value, number=number))

        # -----------------------------------------
        # Attribute 112 - 113 - 114 - 115 - 169 - 171 - 1978 - 1979
        # -----------------------------------------

        # if presence_ht:
        #
        #     for number, valeur in datas_attribut_ht.items():
        #         qs_current.appendRow(self.creation.ligne_attribut(numero=number,
        #                                                     valeur=valeur,
        #                                                     type_attribut=element_type))

        if tag == material_code:
            self.catalog_load(qs_current, child)

        qs_parent.appendRow(qs_list)

    def verif_possibility(self, id_parent: int, type_ele: str) -> bool:

        if id_parent not in self.datas:
            self.datas[id_parent] = type_ele
            return True

        return type_ele == self.datas[id_parent]

    def catalog_load_old(self, qs_parent: QStandardItem, element: etree._Element):

        component_in = element.find("Composant") is not None
        folder_in = element.find("Dossier") is not None

        if folder_in and component_in:
            folder_creation = True
        else:
            folder_creation = False

        new_folder = None

        for child in element:

            tag = child.tag

            if not isinstance(tag, str):
                print("convert_manage -- ConvertExtern -- catalog_load_old -- not isinstance(tag, str)")
                continue

            if tag != "Dossier" and tag != "Composant":
                print(f"convert_manage -- ConvertExtern -- catalog_load_old -- tag is wrong : {tag}")
                continue

            # -----------------------------------------------
            # Folder
            # -----------------------------------------------

            if tag == "Dossier":

                name = child.get('name')

                if not isinstance(name, str):
                    print(f"convert_manage -- ConvertExtern -- catalog_load_old -- folder -- not isinstance(name, str)")
                    continue

                if " - " not in name:
                    description = ""
                else:
                    name, description = name.split(" - ", maxsplit=1)

                qs_list = self.creation.folder_line(value=name,
                                                    description=description,
                                                    tooltips=False)

                qs_current = qs_list[0]

                qs_parent.appendRow(qs_list)

                self.catalog_load_old(qs_parent=qs_current, element=child)

                continue

            # -----------------------------------------------
            # Component
            # -----------------------------------------------

            if tag == "Composant":

                name = child.get('name')

                if not isinstance(name, str):
                    print(f"convert_manage -- ConvertExtern -- catalog_load_old -- Composant -- "
                          f"not isinstance(name, str)")
                    continue

                ele_description = child.find("description")

                if ele_description is None:
                    description = ""
                else:
                    description = ele_description.text

                    if "\n" in description:
                        description = description.replace("\n", " ")

                    if not isinstance(description, str):
                        description = ""

                qs_list = self.creation.component_line(value=name,
                                                       description=description,
                                                       tooltips=False)

                unit_ele = child.find("Unité")

                if unit_ele is None:
                    qs_parent.appendRow(qs_list)
                    continue

                unit = unit_ele.text

                if not isinstance(unit, str):
                    qs_parent.appendRow(qs_list)
                    continue

                qs_current = qs_list[0]

                qs_current.appendRow(self.creation.attribute_line(value=unit,
                                                                  number="202",
                                                                  model_enumeration=self.allplan.model_units))

                if folder_creation:

                    if isinstance(new_folder, Folder):

                        new_folder.appendRow(qs_list)

                    else:

                        qs_folder_list = self.creation.folder_line(value=name,
                                                                   description=description,
                                                                   tooltips=False)

                        new_folder = qs_folder_list[0]

                        qs_parent.appendRow(qs_folder_list)

                        new_folder.appendRow(qs_list)

                else:

                    qs_parent.appendRow(qs_list)


def a___________________export______():
    pass


class ExportExcel(QWidget):

    def __init__(self, allplan: AllplanDatas, model_cat: QStandardItemModel, chemin_fichier: str):
        super().__init__()

        self.allplan = allplan

        self.model_cat = model_cat

        self.chemin_fichier = chemin_fichier
        self.chemin_dossier = find_folder_path(chemin_fichier)

        self.datas_attributs = dict()
        self.dict_index_attributs = dict()
        self.datas = list()

        self.col_id = 3

        self.export_model()

    @staticmethod
    def a___________________attributs______():
        """ Partie réservée à la recherche des données"""
        pass

    def gestion_attributs(self, qs_parent: MyQstandardItem, index_enfant: int):

        qs_val: MyQstandardItem = qs_parent.child(index_enfant, col_cat_value)

        if not isinstance(qs_val, Attribute):
            return None

        qs_numero: MyQstandardItem = qs_parent.child(index_enfant, col_cat_number)
        numero = qs_numero.text()

        if numero == attribute_default_base:
            return ["", ""]

        valeur_attribut: str = qs_val.text()

        type_ele2: str = self.allplan.find_datas_by_number(number=numero, key=code_attr_option)

        if type_ele2 == code_attr_combo_int:

            qs_index: MyQstandardItem = qs_parent.child(index_enfant, col_cat_index)
            valeur_attribut = qs_index.text()

        elif type_ele2 == code_attr_chk:
            if valeur_attribut == "true":
                valeur_attribut = "1"
            else:
                valeur_attribut = "0"

        elif type_ele2 in [code_attr_formule_str, code_attr_formule_int, code_attr_formule_float]:

            if "\n" in valeur_attribut:
                valeur_attribut = valeur_attribut.replace("\n", "")

            try:
                numero_int = int(numero)

                if 1999 < numero_int < 12000:
                    valeur_attribut = "1"

            except ValueError:
                pass

        return [numero, valeur_attribut]

    @staticmethod
    def a___________________model______():
        """ Partie réservée à la recherche des données"""
        pass

    def export_model(self):

        self.datas = [["Type", "Parent", "Attribute", "Value"]]

        self.export_model_creation(self.model_cat.invisibleRootItem())
        self.export_model_to_excel()

    def export_model_creation(self, qs_parent: MyQstandardItem):

        nb_enfant = qs_parent.rowCount()

        if qs_parent == self.model_cat.invisibleRootItem():
            texte_parent = ""
        else:
            texte_parent = qs_parent.text()

        if nb_enfant == 0:
            return

        for index_enfant in range(nb_enfant):
            qs_val: MyQstandardItem = qs_parent.child(index_enfant, col_cat_value)

            texte_val: str = qs_val.text()

            if isinstance(qs_val, Attribute):
                self.export_model_attributs(qs_parent, index_enfant)
                continue

            if isinstance(qs_val, Link):
                self.datas.append(["Link", texte_parent, "", texte_val])
                continue

            if isinstance(qs_val, Folder):
                self.datas.append(["Folder", texte_parent, "", texte_val])

            elif isinstance(qs_val, Material):
                self.datas.append(["Material", texte_parent, "", texte_val])

            elif isinstance(qs_val, Component):
                self.datas.append(["Component", texte_parent, "", texte_val])

            if qs_val.hasChildren():
                self.export_model_creation(qs_val)

    def export_model_attributs(self, qs_parent: MyQstandardItem, index_enfant: int):

        attribut = self.gestion_attributs(qs_parent, index_enfant)

        if not isinstance(attribut, list):
            return

        numero, valeur_attribut = attribut

        texte_parent: str = qs_parent.text()
        self.datas.append(["Attribute", texte_parent, numero, valeur_attribut])

    def export_model_to_excel(self):

        wb = Workbook()
        sheet = wb.active

        try:
            for row in self.datas:
                sheet.append(row)

            wb.save(self.chemin_fichier)

        except Exception as erreur:
            print(f"{erreur}")

            msg(titre=application_title,
                message=self.tr("Une erreur est survenue."),
                icone_avertissement=True,
                details=f"{erreur}")

            return

        msg(titre=application_title,
            message=self.tr("L'export s'est correctement déroulé!"),
            icone_valide=True)

        open_file(self.chemin_fichier)


class ExportCatalog(QObject):

    def __init__(self, catalogue, allplan, file_path: str, brand: str):

        super().__init__()

        self.allplan: AllplanDatas = allplan

        self.catalogue = catalogue

        self.model = self.catalogue.cat_model

        self.file_path = file_path

        self.brand = brand

        self.sauvegarde_catalogue()

    def sauvegarde_catalogue(self):

        tps = time.perf_counter()

        root = etree.Element(self.brand)

        self.sauvegarde_hierarchie(self.model.invisibleRootItem(), root)

        a = self.tr("Une erreur est survenue.")

        try:
            catalogue = etree.tostring(root,
                                       pretty_print=True,
                                       xml_declaration=True,
                                       encoding='UTF-8').decode()

            print(f"ExportCatalog : {time.perf_counter() - tps}s")

        except Exception as erreur:
            msg(titre=application_title,
                message=f'{a} : {self.file_path}',
                icone_critique=True,
                details=f"{erreur}")
            return False

        try:

            with open(self.file_path, "w", encoding="utf_8_sig") as file:
                file.write(catalogue)

        except Exception as erreur:
            msg(titre=application_title,
                message=f'{a} : {self.file_path}',
                icone_critique=True,
                details=f"{erreur}")
            return False

    def sauvegarde_hierarchie(self, qs_parent: MyQstandardItem, racine: etree._Element):

        for index_row in range(qs_parent.rowCount()):

            qs = qs_parent.child(index_row, col_cat_value)

            description = qs_parent.child(index_row, col_cat_desc).text()

            if isinstance(qs, Attribute) or isinstance(qs, Link):
                continue

            if isinstance(qs, Folder):
                node = self.creation_dossier(qs=qs, racine=racine, description=description)

                if not qs.hasChildren():
                    continue

                self.sauvegarde_hierarchie(qs_parent=qs, racine=node)
                continue

            unit = self.get_unit(qs)

            if isinstance(qs, Material):
                # group = self.creation_ouvrage(qs=qs, description=description, unit=unit, node=racine)

                if not qs.hasChildren():
                    continue

                self.sauvegarde_hierarchie(qs_parent=qs, racine=racine)
                continue

            if isinstance(qs, Component):
                self.creation_composant(qs=qs, description=description, unit=unit, group=racine)
                continue

    @staticmethod
    def creation_dossier(qs: QStandardItem, description: str, racine: etree._Element):

        titre = qs.text()
        node = etree.SubElement(racine, "Folder", name=titre, description=description)

        return node

    def creation_ouvrage(self, qs: MyQstandardItem, description: str, unit: str, node: etree._Element):

        group = etree.SubElement(node, "Material", name=qs.text(), description=description, unit=unit)

        self.creation_attributs(qs=qs, definition=group)

        return group

    def creation_composant(self, qs: MyQstandardItem, description: str, unit: str, group: etree._Element):

        position = etree.SubElement(group, "Component", name=qs.text(), description=description, unit=unit)

        self.creation_attributs(qs=qs, definition=position)

    def creation_attributs(self, qs: MyQstandardItem, definition: etree._Element):

        nb_enfants = qs.rowCount()
        plume = True
        trait = True
        couleur = True

        for index_row in range(nb_enfants):

            qstandarditem_enfant_valeur: QStandardItem = qs.child(index_row, col_cat_value)

            if not isinstance(qstandarditem_enfant_valeur, Attribute):
                return

            qstandarditem_enfant_numero: QStandardItem = qs.child(index_row, col_cat_number)
            valeur = qstandarditem_enfant_valeur.text()
            numero = qstandarditem_enfant_numero.text()

            if numero == "207" or numero == "202":
                continue

            if numero in liste_attributs_with_no_val_no_save and valeur == "":
                continue

            if numero == "349":
                plume, trait, couleur = self.gestion_layer(valeur)

            if (numero == "346" and not plume) or (numero == "345" and not trait) or (numero == "347" and not couleur):
                continue

            type_ele2: str = self.allplan.find_datas_by_number(number=numero, key=code_attr_option)

            if type_ele2 == code_attr_combo_int:
                qstandarditem_enfant_combo: QStandardItem = qs.child(index_row, col_cat_index)
                valeur = qstandarditem_enfant_combo.text()
                etree.SubElement(definition, 'Attribute', id=numero, value=valeur)
                continue

            if type_ele2 in [code_attr_formule_str, code_attr_formule_int, code_attr_formule_float]:

                if "\n" in valeur:
                    valeur = valeur.replace("\n", "")

                try:
                    numero_int = int(numero)

                    if 1999 < numero_int < 12000:
                        valeur = "1"

                    else:

                        valeur = self.allplan.convertir_formule(valeur)

                except ValueError:
                    pass

                etree.SubElement(definition, 'Attribute', id=numero, value=valeur)
                continue

            etree.SubElement(definition, 'Attribute', id=numero, value=valeur)
        return

    @staticmethod
    def gestion_layer(numero_style: str):

        plume = True
        trait = True
        couleur = True

        if numero_style == "1":
            plume = False
            return plume, trait, couleur

        if numero_style == "2":
            trait = False
            return plume, trait, couleur

        if numero_style == "3":
            plume = False
            trait = False
            return plume, trait, couleur

        if numero_style == "4":
            couleur = False
            return plume, trait, couleur

        if numero_style == "5":
            plume = False
            couleur = False
            return plume, trait, couleur

        if numero_style == "6":
            trait = False
            couleur = False
            return plume, trait, couleur

        if numero_style == "7":
            plume = False
            trait = False
            couleur = False
            return plume, trait, couleur

        return plume, trait, couleur

    @staticmethod
    def get_unit(qs: MyQstandardItem):

        search = qs.get_attribute_value_by_number("202")

        if search is None:
            return ""

        return search


def a___________________excel______():
    pass


class ObjExcel:

    def __init__(self):
        super().__init__()

        self.name = ""
        self.description = ""
        self.his_parent = None

        self.children_type = None

        self.children = list()
        self.attributes = dict()


class RootExcel(ObjExcel):

    def __init__(self):
        super().__init__()

    def add_children(self, children) -> bool:
        if not isinstance(children, FolderExcel):
            print(f"convertmanage -- ConvertExcel -- RootExcel -- can't add this children in root")
            return False

        self.children.append(children)
        return True

    @staticmethod
    def add_attribute(number: str, value: str):
        return


class FolderExcel(ObjExcel):

    def __init__(self, his_parent, name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number != "207":
            return

        if not isinstance(value, str):
            value = ""

        self.description = value

    def add_children(self, children) -> bool:

        if isinstance(children, FolderExcel):
            if self.children_type == folder_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = folder_code
                self.children.append(children)
                return True

        if isinstance(children, MaterialExcel):
            if self.children_type == material_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = material_code
                self.children.append(children)
                return True

            print("convertmanage -- ConvertExcel -- FolderExcel -- "
                  f"can't add {folder_code} -> {material_code} already exists,")

            return False

        if isinstance(children, ComponentExcel):
            if self.children_type == component_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = component_code
                self.children.append(children)
                return True

            print("convertmanage -- ConvertExcel -- FolderExcel -- "
                  f"can't add {folder_code} -> {component_code} already exists,")
            return False

        print("convertmanage -- ConvertExcel -- FolderExcel -- can't add unknown type,")
        return False


class MaterialExcel(ObjExcel):

    def __init__(self, his_parent: FolderExcel, name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number == "207":

            if not isinstance(value, str):
                value = ""

            self.description = value

            return

        if number in self.attributes:
            return

        self.attributes[number] = value

    def add_children(self, children) -> bool:

        if not isinstance(children, ComponentExcel):
            print(f"convertmanage -- ConvertExcel -- MaterialExcel -- can't add this children in {material_code}")
            return False

        self.children.append(children)
        return True


class ComponentExcel(ObjExcel):

    def __init__(self, his_parent: Union[FolderExcel, MaterialExcel], name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number == "207":

            if not isinstance(value, str):
                value = ""

            self.description = value

            return

        if number in self.attributes:
            return

        self.attributes[number] = value

    @staticmethod
    def add_children(_) -> bool:
        print(f"convertmanage -- ConvertExcel -- Component -- can't add children in {component_code}")
        return False


class ConvertExcel(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan = allplan
        self.creation = self.allplan.creation

        self.file_path = file_path

        self.root = RootExcel()

        self.model_cat = QStandardItemModel()
        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        tps = time.perf_counter()

        workbook = self.excel_load_workbook()

        if not isinstance(workbook, openpyxl.Workbook):
            self.loading_completed.emit(self.model_cat, list(), list())
            return False

        try:

            sheet = workbook.active

            nb_col = sheet.max_column

            if not isinstance(nb_col, int):
                sheet.calculate_dimension(force=True)

                nb_col = sheet.max_column

            if not isinstance(nb_col, int):
                print("convertmanage -- ConvertExcel -- run -- nb_col is None")
                self.run_completed(workbook=workbook)
                return False

            if nb_col == 0:
                print("convertmanage -- ConvertExcel -- run -- max_column_for_row == 0")
                self.run_completed(workbook=workbook)
                return False

            columns_dict = dict()

            for column_index in range(1, nb_col + 1):

                obj_title = sheet.cell(1, column_index)

                if obj_title is None:
                    print("convertmanage -- ConvertExcel -- run -- obj_titre is None")
                    self.run_completed(workbook=workbook)
                    return False

                value = obj_title.value

                if isinstance(value, int):
                    value = f"{value}"

                if not isinstance(value, str):
                    print("convertmanage -- ConvertExcel -- run -- not isinstance(valeur, str)")
                    self.run_completed(workbook=workbook)
                    return False

                value = value.strip().lower()

                if value == "":
                    print("convertmanage -- ConvertExcel -- run -- value empty")
                    break

                columns_dict[value] = column_index - 1

            if len(columns_dict) != 4:
                result = self.load_component_excel(sheet=sheet, columns_dict=columns_dict)

            elif 'type' in columns_dict and 'parent' in columns_dict and 'attribute' in columns_dict \
                    and 'value' in columns_dict:

                result = self.load_material_excel(sheet=sheet, columns_dict=columns_dict)
            else:

                result = self.load_component_excel(sheet=sheet, columns_dict=columns_dict)

            self.run_completed(workbook=workbook)

            print(f"ConvertExtern : {time.perf_counter() - tps}s")

            return result

        except Exception as error:
            print(f"convertmanage -- ConvertExcel -- run -- error : {error}")
            self.loading_completed.emit(self.model_cat, list(), list())
            return False

    def excel_load_workbook(self):

        # -------------
        # Offline
        # -------------

        if not self.file_path.lower().startswith("https"):

            try:
                workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)

                return workbook

            except Exception as error:
                print(f"convert_manage -- ConvertExcel -- excel_load_workbook -- error 2 : {error}")

            return None

        # -------------
        # Online
        # -------------

        try:

            response = requests.get(self.file_path, stream=True, verify=False)

        except Exception as error:
            print(f"convert_manage -- ConvertExcel -- excel_load_workbook -- error : {error}")
            return None

        if not isinstance(response, requests.Response):
            print(f"convert_manage -- ConvertExcel -- excel_load_workbook -- not isinstance(response, Response)")
            return None

        if response.status_code != 200:
            print(f"convert_manage -- ConvertExcel -- excel_load_workbook -- error : {response.status_code}")
            return

        try:
            file_data = BytesIO(response.content)

            workbook = openpyxl.load_workbook(file_data, data_only=True, read_only=True)

            return workbook

        except Exception as error:
            print(f"convert_manage -- ConvertExcel -- excel_load_workbook -- error 2 : {error}")

        return None

    @staticmethod
    def a___________________material______():
        pass

    def load_material_excel(self, sheet, columns_dict: dict) -> bool:

        type_index = columns_dict["type"]
        parent_index = columns_dict["parent"]
        attribute_index = columns_dict["attribute"]
        value_index = columns_dict["value"]

        current_parent = self.root
        current_obj = None

        # -------------------------
        # Read lines
        # -------------------------

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # ----------------------
            # read line
            # ----------------------

            if len(row) < 4:
                print("convertmanage -- ConvertExcel -- load_material_excel -- len(row) < 4")
                continue

            try:

                type_txt: str = row[type_index]
                parent_txt: str = row[parent_index]
                number_txt: str = row[attribute_index]
                value_txt: str = row[value_index]

            except Exception as error:
                print(f"convertmanage -- ConvertExcel -- load_material_excel -- error : {error}")
                continue

            # ----------------------
            # find value
            # ----------------------

            if isinstance(value_txt, int):
                value_txt = f"{value_txt}"

            elif not isinstance(value_txt, str):
                value_txt = ""

            # ----------------------
            # find type element
            # ----------------------

            if not isinstance(type_txt, str):
                print("convertmanage -- ConvertExcel -- load_material_excel -- not isinstance(type_txt, str)")
                continue

            if type_txt == "":
                continue

            if type_txt.lower() in attribute_code_list:

                if not isinstance(current_obj, (FolderExcel, MaterialExcel, ComponentExcel)):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- attribute -- error object")
                    continue

                # ----------------------
                # défintion du numéro
                # ----------------------

                if isinstance(number_txt, int):
                    number_txt = f"{number_txt}"

                elif isinstance(number_txt, str):

                    try:
                        number_txt = str(int(number_txt))

                    except Exception:
                        print("convertmanage -- ConvertExcel -- load_material_excel -- attribute -- "
                              "error convert number")
                        continue

                current_obj.add_attribute(number=number_txt, value=value_txt)
                continue

            # ----------------------
            # Search parent
            # ----------------------

            if isinstance(parent_txt, int):
                parent_txt = ""

            elif not isinstance(parent_txt, str):
                parent_txt = ""

            # Si parent est vide alors root
            if parent_txt == "":
                current_parent = self.root

            elif current_parent is None:
                current_parent = self.root

            elif current_parent.name == parent_txt:

                if isinstance(current_parent, MaterialExcel) and type_txt == material_code:
                    current_parent = current_parent.his_parent

            elif current_parent.name != parent_txt:
                current_parent = self.get_parent(obj_excel=current_parent, parent_name=parent_txt)

                if current_parent is None:
                    print("convertmanage -- ConvertExcel -- load_material_excel -- current_parent is None")
                    continue

            # ----------------------
            # Folder
            # ----------------------

            if type_txt.lower() in folder_code_list:

                obj = FolderExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- folder -- error add_children(obj)")
                    continue

                current_parent = current_obj = obj

            # ----------------------
            # Material
            # ----------------------

            elif type_txt.lower() in material_code_list:

                obj = MaterialExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- material -- error add_children(obj)")
                    continue

                current_parent = current_obj = obj

            # ----------------------
            # Component
            # ----------------------

            elif type_txt.lower() in component_code_list:

                obj = ComponentExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- component -- "
                          "error add_children(obj)")
                    continue

                current_obj = obj

        self.model_add(qs_parent=self.model_cat.invisibleRootItem(), obj_excel=self.root)

        return True

    def get_parent(self, obj_excel: Union[RootExcel, FolderExcel, MaterialExcel, ComponentExcel], parent_name: str):

        if obj_excel is None or obj_excel == self.root:
            print(f"convertmanage -- ConvertExcel -- get_parent -- obj_excel is None")
            return None

        obj_parent = obj_excel.his_parent

        if not isinstance(obj_parent, (RootExcel, FolderExcel, MaterialExcel, ComponentExcel)):
            print(f"convertmanage -- ConvertExcel -- get_parent -- bad obj_parent")
            return None

        if obj_parent.name == parent_name:
            return obj_parent

        children_list = obj_parent.children

        for obj_child in children_list:

            if not isinstance(obj_child, (FolderExcel, MaterialExcel, ComponentExcel)):
                print(f"convertmanage -- ConvertExcel -- get_parent -- bad obj_child")
                continue

            if obj_child.name == parent_name:
                return obj_child

        return self.get_parent(obj_excel=obj_parent, parent_name=parent_name)

    def model_add(self, qs_parent: QStandardItem,
                  obj_excel: Union[RootExcel, FolderExcel, MaterialExcel, ComponentExcel]):

        if not isinstance(obj_excel, (RootExcel, FolderExcel, MaterialExcel, ComponentExcel)):
            print(f"convertmanage -- ConvertExcel -- model_add -- bad object")
            return

        if len(obj_excel.attributes) != 0:

            for number, value in obj_excel.attributes.items():

                model_111 = None

                if number == "111":

                    value_118 = obj_excel.attributes.get("118", 0)

                    if value_118 == "1":
                        model_111 = self.allplan.model_haching
                    elif value_118 == "2":
                        model_111 = self.allplan.model_pattern
                    elif value_118 == "3":
                        model_111 = self.allplan.model_color
                    elif value_118 == "5":
                        model_111 = self.allplan.model_style

                qs_list = self.creation.attribute_line(value=value, number=number, model_enumeration=model_111)

                qs_parent.appendRow(qs_list)

        children = obj_excel.children

        if len(children) == 0:
            return

        for child in children:

            child: ObjExcel

            name = child.name
            description = child.description

            if isinstance(child, FolderExcel):
                qs_list = self.creation.folder_line(value=name, description=description, tooltips=False)

            elif isinstance(child, MaterialExcel):

                qs_list = self.creation.material_line(value=name, description=description, tooltips=False)

            elif isinstance(child, ComponentExcel):
                qs_list = self.creation.component_line(value=name, description=description, tooltips=False)

            else:
                print(f"convertmanage -- ConvertExcel -- model_add -- bad child")
                return

            self.model_add(qs_parent=qs_list[0], obj_excel=child)

            qs_parent.appendRow(qs_list)

    @staticmethod
    def a___________________component______():
        pass

    def load_component_excel(self, sheet, columns_dict: dict) -> bool:

        columns_base = dict()
        columns_other = dict()

        title = find_filename(self.file_path)

        if not isinstance(title, str):
            title = "Excel"

        qs_folder_list = self.creation.folder_line(value=title, tooltips=False)
        root: MyQstandardItem = qs_folder_list[0]

        for value, column_index in columns_dict.items():

            if not isinstance(value, str):
                print("convertmanage -- ConvertExcel -- load_component_excel -- not isinstance(valeur, str)")
                return False

            try:

                int(value)
                number = value

            except Exception:

                number = self.allplan.find_number_by_name(name=value)

                if number == "":
                    print("convertmanage -- ConvertExcel -- load_component_excel -- number not found")
                    continue

            if (number in attribute_val_default_layer or number in attribute_val_default_fill or
                    number in attribute_val_default_room):
                continue

            if number in columns_other:
                print("convertmanage -- ConvertExcel -- load_component_excel -- number in columns_dict")
                continue

            if number == "83" or number == "207":
                columns_base[number] = column_index
            else:
                columns_other[number] = column_index

        if "83" not in columns_base:
            print("convertmanage -- ConvertExcel -- load_component_excel -- 83 not in columns_base")
            return False

        description_in = "207" in columns_base

        # -------------------------
        # Sort header
        # -------------------------

        order_list_columns = {}

        for number in liste_attributs_ordre:
            if number in columns_other:
                order_list_columns[number] = columns_other[number]

        other_columns = [number for number in columns_other if number not in order_list_columns]

        try:

            other_columns.sort(key=int)

        except Exception:
            pass

        for number in other_columns:
            order_list_columns[number] = columns_other[number]

        # -------------------------
        # Read lines
        # -------------------------

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not isinstance(row, tuple):
                print("convertmanage -- ConvertExcel -- load_component_excel -- not isinstance(row, tuple)")
                continue

            column_count = len(row) - 1

            code = row[columns_base["83"]]

            if description_in:
                description = row[columns_base["207"]]
            else:
                description = ""

            qs_component_list = self.creation.component_line(value=code, description=description, tooltips=False)

            qs_component: Component = qs_component_list[0]

            for number, column in order_list_columns.items():

                if column > column_count:
                    print("convertmanage -- ConvertExcel -- load_component_excel -- column > column_count")
                    continue

                value = row[column]

                if isinstance(value, int):
                    value = f"{value}"

                elif isinstance(value, float):
                    value = f"{value:.3f}"

                elif not isinstance(value, str):
                    value = ""

                qs_attribute_list = self.creation.attribute_line(value=value, number=number)

                qs_component.appendRow(qs_attribute_list)

            root.appendRow(qs_component_list)

        self.model_cat.appendRow(qs_folder_list)

        return True

    def run_completed(self, workbook: openpyxl.workbook.Workbook):

        if not isinstance(workbook, openpyxl.workbook.Workbook):
            self.loading_completed.emit(self.model_cat, list(), list())
            return
        try:
            workbook.close()
        except Exception as error:
            print(f"convertmanage -- ConvertExcel -- run_completed -- error: {error}")

        self.loading_completed.emit(self.model_cat, list(), list())

    @staticmethod
    def a___________________end______():
        pass


def a___________________csv______():
    pass


class ConvertCSV(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan = allplan
        self.creation = self.allplan.creation

        self.file_path = file_path

        self.root = RootExcel()

        self.model_cat = QStandardItemModel()
        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        self.separator = "\t"

        self.bdd_type = False

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.bdd_type = msg(titre=application_title,
                            message=self.tr("Définir le type de base de données"),
                            bt_ok=self.tr("Ouvrage"),
                            bt_no=self.tr("Composant"),
                            type_bouton=QMessageBox.Ok | QMessageBox.No,
                            icone_question=True,
                            defaut_bouton=QMessageBox.Ok) == QMessageBox.Ok

        tps = time.perf_counter()

        try:

            lines_list = read_file_to_list(file_path=self.file_path)

            self.model_creation(datas=lines_list)

            self.loading_completed.emit(self.model_cat, list(), list())

            print(f"convert_manage -- ConvertCSV : {time.perf_counter() - tps}s")

            return True

        except Exception as error:
            print(f"convert_manage -- ConvertCSV -- run -- error : {error}")
            self.loading_completed.emit(self.model_cat, list(), list())
            return False

    def model_creation(self, datas: list) -> bool:

        if not isinstance(datas, list):
            print(f"convert_manage -- ConvertCSV -- model_creation -- not isinstance(datas, dict)")
            return False

        if len(datas) < 2:
            print(f"convert_manage -- ConvertCSV -- model_creation -- len(datas) < 2")
            return False

        header = datas[0]
        datas.pop(0)

        if not isinstance(header, str):
            print(f"convert_manage -- ConvertCSV -- model_creation --  not isinstance(header, str)")
            return False

        header = header.strip()

        if self.separator not in header:

            find_separator = False

            for separator in [";", "|"]:
                if separator in header:
                    self.separator = separator
                    find_separator = True
                    break

            if not find_separator:
                print(f"convert_manage -- ConvertCSV -- model_creation -- self.separator not in header")
                return False

        header_list = header.split(self.separator)

        columns_base = dict()
        columns_other = dict()

        for column_index, value in enumerate(header_list):

            if not isinstance(value, str):
                print("convert_manage -- ConvertCSV -- model_creation -- not isinstance(valeur, str)")
                return False

            value = value.strip()

            try:

                int(value)
                number = value

            except Exception:

                number = self.allplan.find_number_by_name(name=value)

                if number == "":
                    print("convert_manage -- ConvertExcel -- load_component_excel -- number not found")
                    continue

            if (number in attribute_val_default_layer or number in attribute_val_default_fill or
                    number in attribute_val_default_room):
                continue

            if number in columns_other:
                print("convert_manage -- ConvertCSV -- model_creation -- number in columns_dict")
                continue

            if number == "83" or number == "207":
                columns_base[number] = column_index
            else:
                columns_other[number] = column_index

        if "83" not in columns_base:
            print("convert_manage -- ConvertCSV -- model_creation -- 83 not in columns_base")
            return False

        description_in = "207" in columns_base

        # -------------------------
        # Sort header
        # -------------------------

        order_list_columns = {}

        for number in liste_attributs_ordre:
            if number in columns_other:
                order_list_columns[number] = columns_other[number]

        other_columns = [number for number in columns_other if number not in order_list_columns]

        other_columns.sort(key=int)

        for number in other_columns:
            order_list_columns[number] = columns_other[number]

        # -------------------------
        # Read lines
        # -------------------------

        title = find_filename(self.file_path)

        if not isinstance(title, str):
            title = "Excel"

        qs_folder_list = self.creation.folder_line(value=title, tooltips=False)
        root: MyQstandardItem = qs_folder_list[0]

        for line in datas:

            if not isinstance(line, str):
                print("convert_manage -- ConvertCSV -- model_creation -- not isinstance(line, str):")
                continue

            line = line.strip()

            if self.separator not in line:
                print(f"convert_manage -- ConvertCSV -- model_creation -- self.separator not in header")
                continue

            line_list = line.split(self.separator)

            column_count = len(line_list)

            code = line_list[columns_base["83"]]

            if description_in:
                description = line_list[columns_base["207"]]
            else:
                description = ""

            if self.bdd_type:

                qs_component_list = self.creation.material_line(value=code, description=description, tooltips=False)

            else:

                qs_component_list = self.creation.component_line(value=code, description=description, tooltips=False)

            qs_component: Component = qs_component_list[0]

            for number, column in order_list_columns.items():

                if column > column_count:
                    print("convert_manage -- ConvertExcel -- load_component_excel -- column > column_count")
                    continue

                value = line_list[column]

                if not isinstance(value, str):
                    value = ""

                qs_attribute_list = self.creation.attribute_line(value=value, number=number)

                qs_component.appendRow(qs_attribute_list)

            root.appendRow(qs_component_list)

        self.model_cat.appendRow(qs_folder_list)

        return True

    @staticmethod
    def a___________________end______():
        pass


def a___________________mxdb______():
    pass


class ConvertMXDB(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        self.allplan = allplan
        self.creation = self.allplan.creation

        self.file_path = file_path

        self.root = RootExcel()

        self.model_cat = QStandardItemModel()
        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        tps = time.perf_counter()

        folder_table = "FLivelliArt"

        folder_id = "ID_Uni"
        folder_cod = "CODLIVART"
        folder_desc1 = "NOMELIVART1"
        folder_sub_id = "PUNTLIVART"

        # -----------------------------------

        component_table = "FElencoPrezzi"

        component_cod = "CODICEARTICOLO_NF"
        component_unit = "UM1"
        component_desc1 = "DESCRIZIONE1"
        component_price = "PREZZO1"

        folder_parent_code = "PUNT_LIVELLOART"

        sub_folder_parent_code = "PUNTARTPADRE"
        sub_folder_id_code = "ID_Uni"

        # -----------------------------------

        folders_dict = dict()
        sub_folders_dict = dict()

        # ----------------------------------- Folder level 0 & Folder level 1 ---------

        try:
            conn = sqlite3.connect(self.file_path)

            cursor = conn.cursor()

            folder_sub_id_index = 0
            folder_cod_index = 1
            folder_desc1_index = 2
            folder_id_index = 3

            query_list = [folder_sub_id, folder_cod, folder_desc1, folder_id]

            query_txt = ", ".join(query_list)

            query = f"""SELECT {query_txt} FROM {folder_table}"""

            cursor.execute(query)

            rows = cursor.fetchall()

            for row in rows:

                id_unique = row[folder_id_index]
                code = row[folder_cod_index]

                if not isinstance(id_unique, int) or not isinstance(code, str):
                    print("Folder -- not isinstance(id_unique, str) or not isinstance(code, str)")
                    continue

                code = code.replace(" ", "")
                code = code.strip()

                if id_unique == "" or code == "":
                    print("Folder -- id_unique or code are empty")
                    continue

                desc = row[folder_desc1_index]

                if not isinstance(desc, str):
                    desc = ""
                else:
                    desc = desc.strip()

                parent_id = row[folder_sub_id_index]

                if not isinstance(parent_id, int):
                    print("Folder -- not isinstance(parent_id, int)")
                    continue

                qs_folder_list = self.creation.folder_line(value=code, description=desc, tooltips=False)

                if parent_id == 0:
                    self.model_cat.appendRow(qs_folder_list)

                    folders_dict[id_unique] = qs_folder_list[0]

                elif parent_id in folders_dict:

                    qs_current = folders_dict[parent_id]

                    qs_current.appendRow(qs_folder_list)

                    folders_dict[id_unique] = qs_folder_list[0]

                else:

                    print("Folder -- parent_id not found")
                    continue

            # ----------------------------------- Folder level 2 ---------

            sub_folder_code_index = 0
            sub_folder_desc_index = 1
            folder_parent_index = 2
            sub_folder_id_index = 3

            query_list = [component_cod, component_desc1, folder_parent_code, sub_folder_id_code]

            query_txt = ", ".join(query_list)

            query = f"""SELECT {query_txt} FROM {component_table} WHERE {sub_folder_parent_code} IS NULL"""

            cursor.execute(query)
            rows = cursor.fetchall()

            for row in rows:

                code = row[sub_folder_code_index]

                if not isinstance(code, str):
                    print("sub_folder -- not isinstance(code, str)")
                    continue

                code = code.strip()

                # -----------

                desc = row[sub_folder_desc_index]

                if not isinstance(desc, str):
                    desc = ""
                else:
                    desc = desc.strip()

                # -----------

                sub_folder_id = row[sub_folder_id_index]

                if not isinstance(sub_folder_id, int):
                    print("sub_folder --  not isinstance(sub_folder_id, int)")
                    continue

                his_parent = row[folder_parent_index]

                if his_parent < 1:
                    continue

                if his_parent not in folders_dict:
                    print("sub_folder -- his_parent not in folders_dict")
                    continue

                qs_current = folders_dict[his_parent]

                qs_folder_list = self.creation.folder_line(value=code, description=desc, tooltips=False)

                qs_current.appendRow(qs_folder_list)

                sub_folders_dict[sub_folder_id] = qs_folder_list[0]

            # -------------------------------

            component_cod_index = 0
            component_unit_index = 1
            component_price_index = 2
            component_desc_index = 3

            sub_folder_id_index = 4

            folder_parent_index = 5

            query_list = [component_cod, component_unit, component_price, component_desc1, sub_folder_parent_code,
                          folder_parent_code]

            query_txt = ", ".join(query_list)

            query = f"""SELECT {query_txt} FROM {component_table} WHERE {sub_folder_parent_code} IS NOT NULL"""

            cursor.execute(query)
            rows = cursor.fetchall()

            for row in rows:

                sub_folder_id = row[sub_folder_id_index]

                his_parent = row[folder_parent_index]

                if not isinstance(sub_folder_id, int) or not isinstance(his_parent, int):
                    print("sub_folder --  not isinstance(sub_folder_id, int) or not isinstance(his_parent, int)")
                    continue

                if sub_folder_id in sub_folders_dict:
                    qs_current = sub_folders_dict[sub_folder_id]

                elif his_parent in folders_dict:
                    print("component -- sub_parent not found")
                    qs_current = folders_dict[his_parent]

                else:
                    print("component -- his_parent not found")
                    continue

                # -----------

                code = row[component_cod_index]

                if not isinstance(code, str):
                    print("component -- not isinstance(code, str)")
                    continue

                code = code.strip()

                # -----------

                desc = row[component_desc_index]

                if not isinstance(desc, str):
                    desc = ""
                else:
                    desc = desc.strip()

                # -----------

                unit = row[component_unit_index]

                if not isinstance(unit, str):
                    unit = ""
                else:
                    unit = unit.strip()

                # -----------

                price = row[component_price_index]

                if not isinstance(price, float):
                    price = "0.00"
                else:
                    price = f"{price:.2f}"

                qs_component_list = self.creation.component_line(value=code, description=desc, tooltips=False)

                qs_component: Component = qs_component_list[0]

                qs_component.appendRow(self.creation.attribute_line(value=unit, number="202"))
                qs_component.appendRow(self.creation.attribute_line(value=price, number="203"))

                qs_current.appendRow(qs_component_list)

            cursor.close()

            self.loading_completed.emit(self.model_cat, list(), list())

            print(f"convertmanage -- ConvertCSV : {time.perf_counter() - tps}s")

            return True

        except sqlite3.Error as error:
            print("Erreur lors de la connexion à SQLite", error)
            return False

    @staticmethod
    def a___________________end______():
        pass


def a___________________nevaris______():
    pass


class ConvertNevarisXml(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)

    def __init__(self, allplan, file_path: str, bdd_title: str, conversion=False):
        super().__init__()

        # --------------
        # Parent
        # --------------

        self.allplan: AllplanDatas = allplan

        self.allplan.creation.attributes_datas.clear()

        self.creation = self.allplan.creation

        # --------------
        # Model
        # --------------

        self.model_cat = QStandardItemModel()

        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description"), "num_attrib"])

        # --------------
        # Variables xml file
        # --------------

        self.file_path = file_path

        self.conversion = conversion

        # --------------
        # Variables xml organization
        # --------------

        self.ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

        self.level_style_dict = {"NEVARIS_STYLE_H1": 0,
                                 "NEVARIS_STYLE_H2": 1,
                                 "NEVARIS_STYLE_H3": 2,
                                 "NEVARIS_STYLE_H4": 3,
                                 "NEVARIS_STYLE_H5": 4,
                                 "NEVARIS_STYLE_H6": 5,
                                 "NEVARIS_STYLE_Default": -1}

        self.code_index = 1
        self.type_index = 2
        self.desc_index = 3
        self.unit_index = 5

        self.root_code = "Raumelement"
        self.element_code = "Element"
        self.component_code = "Position"
        self.link_code = "Link"

        # --------------
        # Variables catalog datas
        # --------------

        self.expanded_list = list()
        self.selected_list = list()

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

        self.errors_list = list()

        # ---------------------------------------
        # LOADING Translation
        # ---------------------------------------

        self.error_material_exist = self.tr("Détection Doublons dans le dossier")
        self.error_renamed = self.tr("a été renommé")

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.model_cat.beginResetModel()

        result = self.parse_excel_xml()

        self.model_cat.endResetModel()

        self.loading_completed.emit(self.model_cat,
                                    self.expanded_list,
                                    self.selected_list)

        return result

    @staticmethod
    def a___________________parse______():
        pass

    def parse_excel_xml(self):

        # --------------
        # Read / Parse file
        # --------------

        try:

            tree = etree.parse(self.file_path)
            root = tree.getroot()

            # --------------
            # Search rows
            # --------------

            # rows = root.xpath('//ss:Worksheet/ss:Table/ss:Row', namespaces=self.ns)
            rows = root.iter("{urn:schemas-microsoft-com:office:spreadsheet}Row")

            # --------------
            # Variables
            # --------------

            rows_list = list()

            cell_datas_previous = dict()

            # --------------
            # Read all rows
            # --------------

            for row in rows:

                # --------------
                # Get style
                # --------------

                style_current = row.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}StyleID', 'Default')

                style_index = self.level_style_dict.get(style_current, -2)

                if style_index == -2:
                    print("convert_manage -- ConvertNevaris -- parse_excel_xml -- style_index not found: "
                          f"({style_current}")
                    continue

                # --------------
                # Get cell datas
                # --------------

                cell_datas = self.get_cells_datas(row=row, style_index=style_index,
                                                  cell_datas_previous=cell_datas_previous)

                if len(cell_datas) == 0:
                    continue

                rows_list.append(cell_datas)

                cell_datas_previous = cell_datas

            # --------------
            # Hierarchy_creation
            # --------------

            qs_levels = {0: self.model_cat.invisibleRootItem()}
            qs_folder = None
            qs_material = None

            qs_folder_dict = dict()

            qs_material_dict = dict()

            for cell_datas in rows_list:

                # --------------
                # Get Datas
                # --------------

                code_current: str = cell_datas["code"]
                type_current: str = cell_datas["type"]
                desc_current: str = cell_datas["desc"]
                unit_current: str = cell_datas["unit"]
                level_index: int = cell_datas["level"]

                # --------------
                # Get parent
                # --------------

                if type_current == folder_code:
                    qs_parent = qs_levels.get(level_index - 1, None)

                elif type_current == material_code:
                    qs_parent = qs_folder

                else:
                    qs_parent = qs_material

                # --------------
                # Check parent
                # --------------

                if not isinstance(qs_parent, QStandardItem):
                    print("convert_manage -- ConvertNevaris -- parse_excel_xml -- "
                          "not isinstance(qs_parent, QStandardItem)")
                    continue

                # --------------
                # Folder creation
                # --------------

                if type_current == folder_code:
                    qs_list = self.creation.folder_line(value=code_current,
                                                        description=desc_current,
                                                        tooltips=self.conversion)

                    qs_parent.appendRow(qs_list)

                    # ----

                    qs_folder = qs_list[0]
                    id_qs = id(qs_folder)

                    # ----

                    qs_levels[level_index] = qs_folder
                    qs_folder_dict[id_qs] = qs_folder

                    continue

                # --------------
                # Material creation
                # --------------

                if type_current == material_code:

                    used_by_links = self.link_list.count(code_current)

                    qs_list = self.creation.material_line(value=code_current,
                                                          description=desc_current,
                                                          used_by_links=used_by_links)

                    # ----

                    qs_material = qs_list[0]

                    # -----

                    id_parent_qs = id(qs_parent)

                    if id_parent_qs not in qs_material_dict:

                        qs_material_dict[id_parent_qs] = [qs_list]

                    else:

                        qs_material_list = qs_material_dict[id_parent_qs]
                        qs_material_list.append(qs_list)

                    continue

                # --------------
                # Component creation
                # --------------

                if type_current == component_code:

                    qs_list = self.creation.component_line(value=code_current, description=desc_current)

                    qs_parent.appendRow(qs_list)

                    # -----

                    if unit_current != "":
                        qs_value = qs_list[0]
                        qs_value.appendRow(self.creation.attribute_line(value=unit_current, number="202"))

                    continue

                # --------------
                # Link creation
                # --------------

                if type_current == link_code:

                    if code_current not in self.material_list:
                        self.errors_list.append(f"Link : {code_current} est orphenlin")
                        continue

                    # -----

                    qs_list = self.creation.link_line(value=code_current, description=desc_current)

                    qs_parent.appendRow(qs_list)

                    continue

                # --------------
                # Error creation
                # --------------

                print("convert_manage -- ConvertNevaris -- parse_excel_xml -- bad type element")

            # --------------
            # Material creation
            # --------------

            for id_folder, qs_material_list in qs_material_dict.items():

                if not isinstance(qs_material_list, list):
                    print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                          "not isinstance(not isinstance(qs_material_list, list)")
                    continue

                qs_parent = qs_folder_dict.get(id_folder, None)

                if not isinstance(qs_parent, Folder):
                    print("convert_manage -- ConvertNevaris -- parce_excel_xml -- not isinstance(qs_parent, Folder)")
                    continue

                parent_row_count = qs_parent.rowCount()

                if parent_row_count != 1:

                    material_row_first = qs_material_list[0]

                    if not isinstance(material_row_first, list):
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                              "not isinstance(material_row_first, list)")
                        continue

                    if len(material_row_first) < 2:
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- len(material_row_first) < 2")
                        continue

                    qs_code: Material = material_row_first[0]
                    qs_desc: Info = material_row_first[1]

                    material_text = qs_code.text()
                    material_desc = qs_desc.text()

                    # -------------

                    qs_list = self.creation.folder_line(value=material_text,
                                                        description=material_desc,
                                                        tooltips=self.conversion)

                    qs_parent = qs_list[0]

                    # -------------

                    qs_parent.appendRow(qs_list)

                for qs_material_row in qs_material_list:

                    if not isinstance(qs_material_row, list):
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                              "not isinstance(qs_material_row, list")
                        continue

                    qs_parent.appendRow(qs_material_row)

            return True

        except Exception as error:
            print(f"convert_manage -- ConvertNevaris -- parse_excel_xml -- error : {error}")
            return False

    @staticmethod
    def a___________________tools______():
        pass

    def get_cells_datas(self, row, style_index: int, cell_datas_previous: dict) -> dict:

        cell_index = 1
        code_current = ""
        type_current = ""
        desc_current = ""

        cell_datas = {"code": "",
                      "type": type_current,
                      "desc": desc_current,
                      "unit": "",
                      "level": style_index}

        for cell in row:

            # --------------
            # Get if new index (cell empty)
            # --------------

            new_index = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Index', None)

            if isinstance(new_index, str):

                try:

                    new_index_int = int(new_index)

                    cell_index = new_index_int

                except Exception:
                    print("convert_manage -- ConvertNevaris -- get_cells_datas -- bad new index")
                    pass

            # --------------
            # Get code
            # --------------

            if cell_index == self.code_index:
                code_current = self.get_datas(cell=cell)
                cell_datas["code"] = code_current

                cell_index += 1
                continue

            # --------------
            # Get type
            # --------------

            if cell_index == self.type_index:

                type_current = self.get_datas(cell=cell)

                type_previous = cell_datas_previous.get("type", folder_code)

                # -------------
                # Folder / Element
                # -------------

                if type_current == self.element_code:
                    cell_index += 1
                    cell_datas["type"] = folder_code
                    continue

                if type_current != self.component_code and type_current != self.link_code:
                    return dict()

                # -------------
                code_previous = cell_datas_previous.get("code", "")
                # -------------

                # -------------
                # Component / Link
                # -------------

                if type_previous == component_code or type_previous == link_code:
                    pass

                # -------------
                # Material
                # -------------

                elif type_previous == folder_code:

                    type_previous = material_code

                    cell_datas_previous["type"] = type_previous

                    # -----------------

                    if code_previous in self.material_list:
                        code_previous_tps = find_new_title(base_title=code_previous,
                                                           titles_list=self.material_upper_list)

                        if code_previous != code_previous_tps:

                            self.errors_list.append(f"Material : {code_previous} a été renommé : {code_previous_tps}")

                            cell_datas_previous["code"] = code_previous_tps

                        else:

                            cell_datas_previous["code"] = code_previous

                    self.material_list.append(code_previous)
                    self.material_list.append(code_previous.upper())

                # -------------
                # Error
                # -------------

                else:
                    print(f"convert_manage -- ConvertNevaris -- get_cells_datas -- bad parent - {type_previous}")
                    return dict()

                # -------------
                # Component
                # -------------

                if type_current == self.component_code:
                    cell_datas["type"] = component_code

                    cell_index += 1
                    continue

                # -------------
                # Link
                # -------------

                cell_datas["type"] = link_code
                self.link_list.append(code_current)

                if code_previous not in self.material_with_link_list:
                    self.material_with_link_list.append(code_previous.upper())

                cell_index += 1
                continue

            # --------------
            # Get description
            # --------------

            if cell_index == self.desc_index:

                desc_current = self.get_datas(cell=cell)
                cell_datas["desc"] = desc_current

                if type_current == self.element_code:

                    if code_current == "":
                        cell_datas["code"] = desc_current

                    return cell_datas

                if type_current == self.root_code:
                    return dict()

                cell_index += 1
                continue

            # --------------
            # Get unit
            # --------------

            if cell_index == self.unit_index:
                cell_datas["unit"] = self.get_datas(cell=cell)

                return cell_datas

            cell_index += 1

        return cell_datas

    def get_datas(self, cell):

        ele = cell.find('ss:Data', namespaces=self.ns)

        if ele is None:
            return

        value = ele.text

        if not isinstance(value, str):
            return ""

        return value

    @staticmethod
    def a___________________end______():
        pass


class ConvertNevarisExcel(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list)
    errors_signal = pyqtSignal(list)

    def __init__(self, allplan, file_path: str, bdd_title: str):
        super().__init__()

        # --------------
        # Parent
        # --------------

        self.allplan: AllplanDatas = allplan

        self.allplan.creation.attributes_datas.clear()

        self.creation = self.allplan.creation

        # --------------
        # Model
        # --------------

        self.model_cat = QStandardItemModel()

        self.model_cat.setHorizontalHeaderLabels([bdd_title, self.tr("Description")])

        # --------------
        # Variables xml file
        # --------------

        self.file_path = file_path
        self.bdd_title = bdd_title

        # --------------
        # Variables xlsx organization
        # --------------

        self.code_index = 0
        self.type_index = 1
        self.desc_index = 2
        self.unit_index = 3
        self.text_index = 4
        self.number_index = 5
        self.formula_index = 8

        self.root_code = "Raumelement"
        self.element_code = "Element"
        self.component_code = "Position"
        self.link_code = "Link"

        self.material_upper_list = list()

        self.errors_list = list()

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:
        self.model_cat.beginResetModel()

        result = self.parse_excel()

        self.model_cat.endResetModel()

        self.loading_completed.emit(self.model_cat, list(), list())

        return result

    @staticmethod
    def a___________________parse______():
        pass

    def parse_excel(self):

        try:
            # --------------
            # Read / Parse file
            # --------------

            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            row_datas_previous = dict()

            row_list = list()

            for row in sheet.iter_rows(min_row=2, values_only=True):

                # ----------------------
                # read line
                # ----------------------

                row_datas = self.get_row_datas(row, row_datas_previous)

                if len(row_datas) == 0:
                    continue

                row_datas_previous = row_datas

                row_list.append(row_datas)

            qs_list = self.creation.folder_line(value=self.bdd_title, description="", tooltips=False)

            qs_root = qs_list[0]

            self.model_cat.appendRow(qs_list)

            qs_material = None

            for row_datas in row_list:

                type_current = row_datas.get("type", self.element_code)

                if type_current == self.element_code:
                    continue

                code_current = row_datas.get("code", "")

                if code_current == "":
                    continue

                desc_current = row_datas.get("desc", "")

                if type_current == material_code:
                    qs_list = self.creation.material_line(value=code_current, description=desc_current)

                    qs_root.appendRow(qs_list)

                    qs_material = qs_list[0]
                    continue

                if type_current != component_code:
                    continue

                qs_list = self.creation.component_line(value=code_current, description=desc_current)

                qs_component = qs_list[0]

                qs_material.appendRow(qs_list)

                attributes_dict = row_datas.get("attributes", list())

                if not isinstance(attributes_dict, dict):
                    return

                for number, value in attributes_dict.items():
                    qs_component.appendRow(self.creation.attribute_line(value=value, number=number))

        except Exception as error:
            print(f"convert_manage -- ConvertNevarisExcel -- parse_excel -- error : {error}")
            return False

        return True

    def get_row_datas(self, row: tuple, last_row: dict) -> dict:

        if len(row) < self.formula_index:
            print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- len(row) < 4")
            return dict()

        try:

            code_previous = last_row.get("code", "")
            type_previous = last_row.get("type", self.element_code)

            # ---------------------
            # Type
            # ---------------------

            type_current = row[self.type_index]

            if not isinstance(type_current, str):
                print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- not isinstance(type_current, str)")
                return dict()

            if type_current == self.root_code:
                return dict()

            elif type_current == self.link_code or type_current == self.component_code:

                # --------------------------------
                # Change last Element to Material
                # --------------------------------

                if type_previous == self.element_code:

                    last_row["type"] = material_code

                    if code_previous.upper() not in self.material_upper_list:

                        self.material_upper_list.append(code_previous.upper())

                    else:

                        code_new = find_new_title(base_title=code_previous, titles_list=self.material_upper_list)

                        last_row["code"] = code_new

                        self.material_upper_list.append(code_new.upper())

                if type_current == self.link_code:
                    return dict()

                if type_current == self.component_code:
                    type_current = component_code

            elif type_current != self.element_code:
                print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- type_current != self.element_code")
                return dict()

            # ---------------------
            # Description
            # ---------------------

            desc_current = row[self.desc_index]

            if not isinstance(desc_current, str):
                desc_current = ""

            # ---------------------
            # code
            # ---------------------

            code_current = row[self.code_index]

            if not isinstance(code_current, str):
                code_current = desc_current

            row_datas = {"code": code_current,
                         "type": type_current,
                         "desc": desc_current}

            if type_current == self.element_code:
                return row_datas

            fusion = code_current == code_previous and type_current == type_previous

            if fusion:

                attributes_dict = last_row.get("attributes", dict())

            else:

                attributes_dict = dict()

            # ---------------------
            # Unit
            # ---------------------

            unit_current = row[self.unit_index]

            if not isinstance(unit_current, str):
                unit_current = ""

            if unit_current != "" and "202" not in attributes_dict:
                attributes_dict["202"] = unit_current

            # ---------------------
            # Text
            # ---------------------

            text_current: str = row[self.text_index]

            if not isinstance(text_current, str):
                text_current = ""

            if text_current != "" and "208" not in attributes_dict:
                attributes_dict["208"] = text_current

            # ---------------------
            # Number
            # ---------------------

            number_current: str = row[self.number_index]

            if not isinstance(number_current, str):
                number_current = ""

            # ---------------------
            # Formula
            # ---------------------

            formula_current = row[self.formula_index]

            if not isinstance(formula_current, str):
                formula_current = ""

            if number_current != "" and number_current not in attributes_dict:
                attributes_dict[number_current] = formula_current

            # ---------------------

            if not fusion:
                row_datas["attributes"] = attributes_dict

                return row_datas

            return dict()

        except Exception as error:
            print(f"convert_manage -- ConvertNevarisExcel -- get_row_datas --  error : {error}")
            return dict()


def a___________________end______():
    pass
